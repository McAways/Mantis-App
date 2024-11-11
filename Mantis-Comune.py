import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter import StringVar
import openpyxl
from openpyxl import load_workbook
import pandas as pd
import requests
import json
from datetime import datetime, timedelta, time
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Color, Alignment, PatternFill, Border, Side
import math
import re

# Variáveis globais para armazenar credenciais após seleção
dados_selecionados = {}

def salvar_dataframe(df, nome):
    output_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel file", "*.xlsx")],
        initialfile=f'{nome}.xlsx'
    )
    if output_path:
        df.to_excel(output_path, index=False)
        exibir_log(f"Arquivo {nome} salvo com Sucesso")

def aplicar_estilo(output_path):
    # Carregar a planilha salva
    wb = load_workbook(output_path)
    ws = wb.active

    # Ajuste da largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Pega a letra da coluna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Preenchimento de fundo
    fill_grey = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Estilo de alinhamento e preenchimento das células
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Começa na segunda linha
        for cell in row:
            if cell.row % 2 == 0:
                cell.fill = fill_grey  # Linhas pares recebem fundo cinza
            else:
                cell.fill = fill_white  # Linhas ímpares recebem fundo branco
            cell.alignment = Alignment(horizontal="left", vertical="center")  # Alinhamento das células

    # Aplicar bordas a todas as células
    apply_borders(ws)

    # Salvar a planilha com as formatações
    wb.save(output_path)

def abrir_popup_selecao_coleta():
    popup = ttk.Toplevel(root)
    popup.title("Selecione as funções para coletar")
    popup.geometry("300x200")
    
    frame_popup = ttk.Frame(popup)
    frame_popup.pack(expand=True)

    # Variáveis para armazenar as seleções
    selecoes = {
        "Coletar Plan. Cadastro": [tk.BooleanVar(), coleta_empresa],
        "Coletar Justificativas": [tk.BooleanVar(), coleta_justificativa],
        "Coletar Pessoas Para Alt.": [tk.BooleanVar(), alteracao_pessoas],
        "Coleta de Planilha de Justificativa": [tk.BooleanVar(), coleta_de_planilhas_justificativa],
        "Planilha de Marcação": [tk.BooleanVar(), coleta_planilha_marcacoes],
        "Planilha de Inconsistência": [tk.BooleanVar(), coleta_planilha_marcacoes_inconsistencia],
        "Planilha de Incomum": [tk.BooleanVar(), coleta_planilha_marcacoes_incomum]
    }

    for i, (nome, (var, _)) in enumerate(selecoes.items()):
        ttk.Checkbutton(frame_popup, text=nome, variable=var).grid(row=i, column=0, sticky="w", padx=10, pady=5)

    # Centraliza o botão abaixo dos Checkbuttons
    ttk.Button(
        frame_popup, text="Iniciar Coleta", 
        command=lambda: iniciar_coleta(selecoes, popup)
    ).grid(row=len(selecoes), column=0, pady=10)
    
    frame_popup.grid_rowconfigure(len(selecoes), weight=1)
    frame_popup.grid_columnconfigure(0, weight=1)

def envio_justificativa():
    """Função para enviar marcações para a API usando o Excel selecionado."""
    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione o arquivo de marcações",
        filetypes=[("Arquivo Excel", "*.xlsx *.xls")]
    )
    if not caminho_arquivo:
        return

    try:
        df = pd.read_excel(caminho_arquivo)
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao ler o arquivo: {e}")
        return
    
    
    url = 'https://www.mdcomune.com.br/RestServiceApi/PreJustificationRequest/PreJustificationRequest'
    
    resultados = []

    for index, row in df.iterrows():
        idjust = row['Justificativa']
        idfunc = row['Id Funcionário']
        # Converte Timestamp para string no formato YYYY-MM-DD
        data = row['Data'].strftime('%Y-%m-%d')
        # Converte time para string no formato HH:MM
        horas = row['Hora'].strftime('%H:%M')

        headers = {
            "identifier": dados_selecionados["CNPJ"],
            "key": dados_selecionados["Chave API"],
            'User-Agent': 'PostmanRuntime/7.30.0'
        }

        payload = {
            "IdJustification": idjust,
            "IdUser": "1",
            "IdEmployee": idfunc,
            "QtdHours": horas,
            "Date": data,
            "Notes": "Enviado via API",
            "RequestType": "1",
            "ResponseType": "AS400V1"
        }

        response = requests.post(url,json=payload, headers=headers
        )
        
        try:
            response_json = response.json()
            status = "Sucesso" if response_json.get("Sucesso") else "Falha"
            mensagem = response_json.get("Mensagem", "Erro desconhecido")
        except ValueError:
            status = "Falha"
            mensagem = "Resposta inválida da API"

        resultados.append([idfunc, data, status, mensagem])
        exibir_log(f"Matricula: {idfunc} | Status: {status} | Mensagem: {mensagem}")

    # Exibe mensagem final
    messagebox.showinfo("Processo Concluído", "Envio de marcações concluído!")

def abrir_popup_selecao_pessoas():
    popup = ttk.Toplevel(root)
    popup.title("Selecione as funções para coletar")
    popup.geometry("300x200")
    
    frame_popup = ttk.Frame(popup)
    frame_popup.pack(expand=True)

    # Variáveis para armazenar as seleções
    selecoes = {
        "Cadastrar Pessoas": [tk.BooleanVar(), cadastrar_pessoas],
        "Alteração de Pessoas": [tk.BooleanVar(), alteracao_pessoas_envio],
        "Envio Justificativa": [tk.BooleanVar(), envio_justificativa]
    }

    # Criação dinâmica dos Checkbuttons
    for i, (nome, (var, _)) in enumerate(selecoes.items()):
        ttk.Checkbutton(frame_popup, text=nome, variable=var).grid(row=i, column=0, sticky="w", padx=10, pady=5)

    # Centraliza o botão abaixo dos Checkbuttons
    ttk.Button(
        frame_popup, text="Iniciar Coleta", 
        command=lambda: iniciar_coleta(selecoes, popup)
    ).grid(row=len(selecoes), column=0, pady=10)
    
    frame_popup.grid_rowconfigure(len(selecoes), weight=1)
    frame_popup.grid_columnconfigure(0, weight=1)

def iniciar_coleta(selecoes, popup):
    """Inicia a coleta com base nas seleções."""
    popup.destroy()  # Fecha o popup

    for nome, (var, func) in selecoes.items():
        if var.get():  # Se a opção foi marcada
            exibir_log(f"Iniciando: {nome}")
            func()  # Chama a função correspondente

    messagebox.showinfo("Concluido", "Função Executada com Sucesso")

def coleta_de_planilhas_justificativa ():
    
    df2 = pd.DataFrame(columns=('Justificativa', 'Id Funcionário', 'Data', "Hora"
    ))
    

    output_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",  # extensão padrão
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],  # tipos de arquivo
        title="Salvar arquivo de marcações como"
    )
    
    output_path1 = filedialog.asksaveasfilename(
        defaultextension=".xlsx",  # extensão padrão
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],  # tipos de arquivo
        title="Salvar arquivo de justificativas como"
    )
    
    if output_path1:  # Verifica se um caminho foi selecionado
        df2.to_excel(output_path1, index=False)

        # Ajustando a largura das colunas após salvar
        wb = load_workbook(output_path1)
        ws = wb.active
        
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]  # Transforma a coluna em uma lista
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)  # +2 para um pouco de espaçamento
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
        wb.save(output_path1)
        wb.close()
    
    exibir_log(f'Planilhas salvas em {output_path}')

def apply_borders(ws):
    # Define o estilo da borda
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Aplica a borda a todas as células do worksheet
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

def generate_date_range(start_date, end_date):
    return pd.date_range(start=start_date, end=end_date)

def clean_json_date(date_str):
        # Remove o dia da semana (últimos 3 caracteres)
    return date_str[:10]

def filtra_marcacoes_impares_e_htrab_vazio(entradas):
    marcacoes_impares = []

    marcacoes_filtradas = []

    for entrada in entradas:
        apontamentos = entrada.get("Apontamentos", "").strip()
        htrab = entrada.get("HTrab", "").strip()
        
        # Verifica se o campo HTrab está vazio e se o número de apontamentos é ímpar
        if not htrab and apontamentos:
            # Divide os apontamentos usando o espaço como delimitador
            intervals = apontamentos.split()
            
            # Verifica se o número de intervalos é ímpar
            if len(intervals) % 2 != 0:
                marcacoes_filtradas.append(entrada)

    return marcacoes_filtradas

def coleta_planilha_marcacoes_inconsistencia():
    output_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                    title="Salvar arquivo como"
                    )
    
    url = 'https://www.mdcomune.com.br/RestServiceApi/ReportEmployeePunch/GetReportEmployeePunch'
    
    headers = {
            "identifier": dados_selecionados["CNPJ"],
            "key": dados_selecionados["Chave API"],
            'User-Agent': 'PostmanRuntime/7.30.0'
    }

    payload = {
            "MatriculaPessoa": [],
            "DataInicio": dados_selecionados["Data Início"],
            "DataFim": dados_selecionados["Data fim"],
            "ResponseType": "AS400V1"
    }
    
    response = requests.post(url, json=payload, headers=headers)

    if response.status_code == 200:
        try:
            data = response.json()
            if "Obj" in data and isinstance(data['Obj'], list):
                all_entries = []
                
                for item in data['Obj']:
                    fixed_data = {
                        'Empresa': item['InfoEmpresa']['Nome'],
                        'Funcionario': item['InfoFuncionario']['Nome'],
                        'Matricula': item['InfoFuncionario']['Matricula'],
                        'Estrutura': item['InfoFuncionario']['Estrutura']
                    }

                    for entrada in item['Entradas']:
                        data_limpa = clean_json_date(entrada['Data'])
                        entrada_data_formatada = datetime.strptime(data_limpa, "%d/%m/%Y")

                        # Substituímos a chamada para filtra_marcacoes_impares pela nova função
                        marcacoes_impares = filtra_marcacoes_impares_e_htrab_vazio([entrada])
                        if marcacoes_impares:
                            for marcacao in marcacoes_impares:
                                entry_data = {
                                    'Data': entrada_data_formatada.strftime("%d/%m/%Y"),
                                    'Horario': marcacao['Horario'],
                                    'Apontamentos': marcacao['Apontamentos']
                                }
                                # Combina dados fixos e dados filtrados e os adiciona ao DataFrame
                                combined_data = {**fixed_data, **entry_data}
                                all_entries.append(combined_data)

                # Verifica se há entradas antes de criar o DataFrame
                if all_entries:
                    final_df = pd.DataFrame(all_entries)

                    final_df['Entrada'] = ''
                    final_df['Almoço Ida'] = ''
                    final_df['Almoço Volta'] = ''
                    final_df['Saida'] = ''
                    
                    final_df.to_excel(output_path, index=False)
                    
                    wb = load_workbook(output_path)
                    ws = wb.active

                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        ws.column_dimensions[column].width = adjusted_width

                    fill_grey = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        for cell in row:
                            if cell.row % 2 == 0:
                                cell.fill = fill_grey
                            else:
                                cell.fill = fill_white
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                            
                    apply_borders(ws)
                    wb.save(output_path)
                    print(f"Arquivo Excel salvo com sucesso em {output_path}")
                else:
                    print("Nenhuma marcação ímpar com HTrab vazio encontrada no período.")
            else:
                print("Nenhum dado no campo 'Obj'.")
        except ValueError as e:
            print(f"Erro ao decodificar JSON: {e}")
            print("Conteúdo da resposta:")
            print(response.text)
    else:
        print(f"Falha: {response.status_code}")
        print(response.text)

    exibir_log(f'Planilha de Marcação salva em {output_path}')
    print(payload)

def process_incomum(df):
    incomuns = pd.DataFrame(columns=df.columns)

    for index, row in df.iterrows():
        apontamentos = row['Apontamentos']
        apontamentos_parts = re.findall(r'\d{2}:\d{2}', str(apontamentos))

        horario = row['Horario']
        horario_parts = re.findall(r'\d{2}:\d{2}', str(horario))

        if len(apontamentos_parts) > 0 and len(horario_parts) > 0:
            apontamentos_length = len(apontamentos_parts)
            horario_length = len(horario_parts)

            # Verifica se o número de intervalos é par e diferente do número de horários
            if apontamentos_length % 2 == 0 and apontamentos_length != horario_length:
                incomuns = pd.concat([incomuns, row.to_frame().T])

    # Formata a coluna de data
    incomuns['Data'] = pd.to_datetime(incomuns['Data'], format='%d/%m/%Y').dt.strftime('%d/%m/%Y')
    return incomuns

def coleta_planilha_marcacoes_incomum():
    output_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                    title="Salvar arquivo como"
                    )
    
    url = 'https://www.mdcomune.com.br/RestServiceApi/ReportEmployeePunch/GetReportEmployeePunch'
    
    headers = {
            "identifier": dados_selecionados["CNPJ"],
            "key": dados_selecionados["Chave API"],
            'User-Agent': 'PostmanRuntime/7.30.0'
    }

    payload = {
            "MatriculaPessoa": [],
            "DataInicio": dados_selecionados["Data Início"],
            "DataFim": dados_selecionados["Data fim"],
            "ResponseType": "AS400V1"
    }
    
    response = requests.post(url, json=payload, headers=headers)

    if response.status_code == 200:
        try:
            data = response.json()
            if "Obj" in data and isinstance(data['Obj'], list):
                all_entries = []
                
                for item in data['Obj']:
                    fixed_data = {
                        'Empresa': item['InfoEmpresa']['Nome'],
                        'Funcionario': item['InfoFuncionario']['Nome'],
                        'Matricula': item['InfoFuncionario']['Matricula'],
                        'Estrutura': item['InfoFuncionario']['Estrutura']
                    }

                    for entrada in item['Entradas']:
                        data_limpa = clean_json_date(entrada['Data'])
                        entrada_data_formatada = datetime.strptime(data_limpa, "%d/%m/%Y")

                        entry_data = {
                            'Data': entrada_data_formatada.strftime("%d/%m/%Y"),
                            'Horario': entrada['Horario'],
                            'Apontamentos': entrada['Apontamentos']
                        }
                        # Adiciona os dados fixos e dados da marcação
                        combined_data = {**fixed_data, **entry_data}
                        all_entries.append(combined_data)

                # Verifica se há entradas antes de criar o DataFrame
                if all_entries:
                    final_df = pd.DataFrame(all_entries)
                    
                    final_df['Entrada'] = ''
                    final_df['Almoço Ida'] = ''
                    final_df['Almoço Volta'] = ''
                    final_df['Saida'] = ''

                    # Processa as marcações incomuns
                    incomuns_df = process_incomum(final_df)

                    # Se houver marcações incomuns, salva elas no arquivo
                    if not incomuns_df.empty:
                        # Salvar o DataFrame com as marcações incomuns em Excel
                        incomuns_df.to_excel(output_path, index=False)

                        # Aplicar o estilo no arquivo salvo
                        aplicar_estilo(output_path)

                    # Caso não haja marcações incomuns, informa ao usuário
                    else:
                        print("Não há marcações incomuns para salvar.")
                    
                else:
                    print("Nenhuma marcação válida encontrada no período.")
            else:
                print("Nenhum dado no campo 'Obj'.")
        except ValueError as e:
            print(f"Erro ao decodificar JSON: {e}")
            print("Conteúdo da resposta:")
            print(response.text)
    else:
        print(f"Falha: {response.status_code}")
        print(response.text)

    exibir_log(f'Planilha de Marcação (incomum) salva em {output_path}')
    print(payload)
    
def coleta_planilha_marcacoes():
    
    output_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",  # extensão padrão
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],  # tipos de arquivo
                  title="Salvar arquivo como"
                    )
    
    # URL do endpoint e payload
    url = 'https://www.mdcomune.com.br/RestServiceApi/ReportEmployeePunch/GetReportEmployeePunch'  # Modifique para o endpoint correto
    
    # Headers da requisição
    headers = {
            "identifier": dados_selecionados["CNPJ"],
            "key": dados_selecionados["Chave API"],
            'User-Agent': 'PostmanRuntime/7.30.0'
    }

    payload = {
            "MatriculaPessoa": [],
            "DataInicio": dados_selecionados["Data Início"],
            "DataFim": dados_selecionados["Data fim"],
            "ResponseType":"AS400V1"
    }
    
    response = requests.post(url,json=payload,headers=headers)

    if response.status_code == 200:
        try:
            data = response.json()
            if "Obj" in data and isinstance(data['Obj'], list):
                all_entries = []
                all_fixed_data = []

                # Defina o intervalo de datas do relatório
                start_date = datetime.strptime(payload["DataInicio"], "%d/%m/%Y")
                end_date = datetime.strptime(payload["DataFim"], "%d/%m/%Y")
                full_date_range = generate_date_range(start_date, end_date)

                # Itera sobre cada objeto no campo "Obj"
                for item in data['Obj']:
                    # Adiciona os dados fixos para cada funcionário
                    fixed_data = {
                        'Empresa': item['InfoEmpresa']['Nome'],
                        'Funcionario': item['InfoFuncionario']['Nome'],
                        'Matricula': item['InfoFuncionario']['Matricula'],
                        'Estrutura': item['InfoFuncionario']['Estrutura']
                    }

                    # Cria um dicionário para armazenar as marcações por data
                    entradas_por_data = {}
                    for entrada in item['Entradas']:
                        # Limpa a data removendo o dia da semana
                        data_limpa = clean_json_date(entrada['Data'])
                        # Converte a data para o formato datetime
                        entrada_data_formatada = datetime.strptime(data_limpa, "%d/%m/%Y")
                        entradas_por_data[entrada_data_formatada] = entrada

                    # Para cada data no intervalo, insere os dados ou deixa em branco
                    for date in full_date_range:
                        if date in entradas_por_data:
                            entrada = entradas_por_data[date]
                            entry_data = {
                                'Data': date.strftime("%d/%m/%Y"),
                                'Horario': entrada['Horario'],
                                'Apontamentos': entrada['Apontamentos']
                            }
                        else:
                            # Preenche com valores em branco para datas sem entrada
                            entry_data = {
                                'Data': date.strftime("%d/%m/%Y"),
                                'Horario': '',
                                'Apontamentos': ''
                            }

                        # Adiciona os dados fixos e de entrada combinados
                        combined_data = {**fixed_data, **entry_data}
                        all_entries.append(combined_data)

                # Converte a lista de dados em um DataFrame
                final_df = pd.DataFrame(all_entries)

                # Adiciona a nova coluna "Justificativas" com valores em branco
                final_df['Entrada'] = ''
                final_df['Almoço Ida'] = ''
                final_df['Almoço Volta'] = ''
                final_df['Saida'] = ''
                
                final_df.to_excel(output_path, index=False)
                
                # Cria um Workbook do openpyxl
                wb = openpyxl.load_workbook(output_path)
                ws = wb.active

                # Ajusta a largura das colunas automaticamente
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width

                # Formatação de cor alternada nas linhas
                fill_grey = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        # Alterna entre as cores
                        if cell.row % 2 == 0:
                            cell.fill = fill_grey
                        else:
                            cell.fill = fill_white
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        
                apply_borders(ws)

                # Salva o arquivo Excel com formatação
                wb.save(output_path)
                print(f"Arquivo Excel salvo com sucesso em {output_path}")
            else:
                print("Nenhum dado no campo 'Obj'.")
        except ValueError as e:
            print(f"Erro ao decodificar JSON: {e}")
            print("Conteúdo da resposta:")
            print(response.text)
    else:
        print(f"Falha: {response.status_code}")
        print(response.text)
    
    exibir_log(f'PLanilha de Marcação salva em {output_path}')
    print(payload)

def get_data_from_api(url, payload, headers):
    """Faz uma requisição para a API e transforma a resposta em DataFrame."""
    try:
        response = requests.post(url, json=payload, headers=headers)
        response.raise_for_status()  # Verifica se a requisição foi bem-sucedida

        data = response.json()  # Decodifica o JSON

        # Verifica se 'Obj' é uma lista e normaliza diretamente
        if 'Obj' in data and isinstance(data['Obj'], list):
            obj_data = pd.json_normalize(data['Obj'])
            return obj_data

        # Caso 'Obj' seja uma string JSON, faz o parsing
        elif 'Obj' in data and isinstance(data['Obj'], str):
            obj_data = json.loads(data['Obj'])  # Converte a string JSON em lista
            return pd.json_normalize(obj_data)

        # Normaliza o JSON principal se 'Obj' não estiver presente
        else:
            return pd.json_normalize(data)

    except ValueError as e:
        print(f"Erro ao decodificar JSON: {e}")
        print("Conteúdo da resposta:", response.text)
    except requests.exceptions.RequestException as e:
        print(f"Erro na requisição: {e}")

def preencher_detalhes():
    """Preenche os detalhes da empresa selecionada na interface."""
    empresa_selecionada = razao_social_var.get()
    empresa_detalhes = df_empresas[df_empresas["Razão Social"] == empresa_selecionada].iloc[0]

    # Preenche os campos com os dados da empresa selecionada
    cnpj_var.set(empresa_detalhes["CNPJ"])
    chave_var.set(empresa_detalhes["Chave API"])
    cpf_var.set(empresa_detalhes.get("CPF Responsável", ""))  # CPF pode ser opcional

def combinar_data_hora(data_completa, hora, tipo_marcacao=""):
    # Verifica se ambos os valores estão disponíveis
    if pd.notna(data_completa) and pd.notna(hora):
        # Garante que data_completa esteja no formato string
        if isinstance(data_completa, pd.Timestamp):
            data = data_completa.strftime("%d/%m/%Y")
        else:
            data = str(data_completa)
        
        # Converte 'hora' para string sem segundos, se necessário
        hora_str = hora if isinstance(hora, str) else hora.strftime("%H:%M")
        if len(hora_str) > 5:  # Caso tenha segundos, remove-os
            hora_str = hora_str[:5]
        
        # Combina a data e hora no formato especificado
        data_hora_str = f"{data} {hora_str}"
        try:
            # Tenta converter no formato %H:%M
            data_hora_dt = datetime.strptime(data_hora_str, "%d/%m/%Y %H:%M")
            
            # Ajusta a data para o próximo dia se for uma marcação noturna
            if tipo_marcacao != "Entrada":
                if data_hora_dt.hour == 23 and data_hora_dt.minute == 59 or data_hora_dt.hour < 6:
                    data_hora_dt += timedelta(days=1)
                    
            return data_hora_dt.strftime("%d/%m/%Y %H:%M")
        
        except ValueError:
            try:
                # Caso tenha segundos, tenta remover e converter novamente
                data_hora_dt = datetime.strptime(data_hora_str, "%d/%m/%Y %H:%M:%S")
                
                # Ajusta a data para o próximo dia se for uma marcação noturna
                if tipo_marcacao != "Entrada":
                    if data_hora_dt.hour == 23 and data_hora_dt.minute == 59 or data_hora_dt.hour < 6:
                        data_hora_dt += timedelta(days=1)
                        
                return data_hora_dt.strftime("%d/%m/%Y %H:%M")
            
            except ValueError as e:
                print(f"Erro ao converter data e hora: {e}")
                return None
    return None

def enviar_marcacoes():
    """Função para enviar marcações para a API usando o Excel selecionado."""
    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione o arquivo de marcações",
        filetypes=[("Arquivo Excel", "*.xlsx *.xls")]
    )
    if not caminho_arquivo:
        return

    try:
        df = pd.read_excel(caminho_arquivo)
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao ler o arquivo: {e}")
        return

    # URL e Headers para a API
    url = "https://www.mdcomune.com.br/RestServiceApi/Mark/SetMarks"
    Headers = {
        "identifier": dados_selecionados['CNPJ'],
        "key": dados_selecionados['Chave API'],
        "cpf": dados_selecionados['CPF Responsável'],
        'User-Agent': 'PostmanRuntime/7.30.0'
    }
    
    resultados = []
    
    for index, row in df.iterrows():
        matricula = row['Matricula']
        data_completa = row['Data']
        
        # Lista de marcações para enviar, com seus respectivos tipos
        marcacoes = {
            "Entrada": row['Entrada'],
            "Almoço Ida": row['Almoço Ida'],
            "Almoço Volta": row['Almoço Volta'],
            "Saida": row['Saida']
        }
        
        for tipo, hora in marcacoes.items():
            if pd.notna(hora):  # Verifica se a marcação está preenchida
                data_hora_marcacao = combinar_data_hora(data_completa, hora, tipo)
                if data_hora_marcacao:
                    payload = {
                        "Matricula": matricula,
                        "DataHoraApontamento": data_hora_marcacao,
                        "CpfResponsavel": dados_selecionados['CPF Responsável'],
                        "ResponseType": "AS400V1"
                    }
                    # Envia a requisição para o endpoint
                    response = requests.post(url, json=payload, headers=Headers)

                    # Define valores padrão para status e mensagem
                    status = "Falha"
                    mensagem = "Resposta inválida da API"

                    try:
                        response_json = response.json()
                        status = "Sucesso" if response_json.get("Sucesso") else "Falha"
                        mensagem = response_json.get("Mensagem", "Erro desconhecido")
                        print(f"{tipo} enviada para {matricula}: {status} - {mensagem}")
                        print(payload)
                    except ValueError:
                        print(f"{tipo} enviada para {matricula}: {status} - {mensagem}")
                        print(payload)

                    resultados.append([matricula, data_completa, mensagem])
                    exibir_log(f"Matricula: {matricula} | Status: {status} | Mensagem: {mensagem}")

def cadastrar_pessoas():
    """Cadastra pessoas na API utilizando dados do Excel."""
    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione o arquivo de cadastro de pessoas",
        filetypes=[("Arquivo Excel", "*.xlsx *.xls")]
    )
    if not caminho_arquivo:
        return

    try:
        df = pd.read_excel(caminho_arquivo)
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao ler o arquivo: {e}")
        return
    resultados = []
    url = "https://www.mdcomune.com.br/RestServiceApi/People/SavePerson"
    headers = {
        "identifier": dados_selecionados["CNPJ"],
        "key": dados_selecionados["Chave API"],
        "cpf": dados_selecionados['CPF Responsável'],
        "User-Agent": "PostmanRuntime/7.30.0"
    }

    # Iteração sobre cada pessoa no DataFrame e envio da requisição POST
    for _, row in df.iterrows():
        try:
            Matricula = row['Matricula']
            Cracha = row['Cracha']
            Nome = row['Nome Completo']
            Cpf = row['CPF']
            DataAdm = pd.to_datetime(row["Admissão"]).strftime('%d-%m-%Y')
            DataNasc = pd.to_datetime(row["Nascimento"]).strftime('%d-%m-%Y')
            BaseHoras = row['Base de Horas']
            EstruturaOrg = row['Estrutura']
            Horario = row['Horário']
            RegraCalculo = row['Cálculo']
            Pis = row.get('PIS', None)
            PisAuto = row.get('Possui PIS?', 1)
            Sexo = row['Sexo']
            Cargo = str(row.get('Cargo', None))
            
            payload = {
                "Matricula": Matricula,
                "Cracha": Cracha,
                "Nome": Nome,
                "Cpf": Cpf,
                "CpfResponsavel": dados_selecionados["CPF Responsável"],
                "DataAdmissao": DataAdm,
                "DataNascimento": DataNasc,
                "BaseHoras": BaseHoras,
                "Estrutura": {"Id": EstruturaOrg},
                "TipoFuncionario": {"IdTipoFuncionario": 1},
                "TipoSalario": {"Id": 101},
                "Horarios": [
                    {
                        "Id": 0,
                        "Horario": {"Id": Horario},
                        "Inicio": DataAdm,
                        "Fim": "31/12/9999 00:00:00"
                    }
                ],
                "RegrasCalculo": [
                    {
                        "Id": 0,
                        "Regra": {"Id": RegraCalculo},
                        "Inicio": DataAdm,
                        "Fim": "31/12/9999 00:00:00"
                    }
                ],
                "FlagGerarNumeroPISAutomatico": bool(PisAuto == 0),
                "Sexo": Sexo
            }
            
            if pd.notna(Cargo):
                payload['Cargo'] = {"Id":Cargo}

            # Adicionar PIS se necessário
            if PisAuto == 1 and Pis:
                payload["CodigoPis"] = Pis

            # Enviar a requisição
            response = requests.post(url, json=payload, headers=headers)
            response_json = response.json()

            # Verificar sucesso ou falha
            if response_json.get("Sucesso"):
                status = "Sucesso"
                mensagem = response_json.get("Mensagem", "")
            else:
                status = "Falha"
                mensagem = response_json.get("Mensagem", "Erro desconhecido")

        except Exception as e:
            status = "Erro"
            mensagem = str(e)

        # Armazenar o resultado
        resultados.append([Matricula, status, mensagem])
        
        exibir_log(f"Matricula: {Matricula} | Status: {status} | Mensagem: {mensagem}")
        
def coleta_justificativa():
    url1 = 'https://www.mdcomune.com.br/RestServiceApi/People/SearchPeople'
    url2 = 'https://www.mdcomune.com.br/RestServiceApi/Justification/GetJustification'
    
    payload1 = {"Matricula": 0}
    payload2 = {"Code": 0,
            "IdType": 1202,
            "ResponseType": "AS400V1"}  # Modifique conforme necessário
    
    headers = {
        "identifier": dados_selecionados['CNPJ'],
        "key": dados_selecionados['Chave API'],
        'User-Agent': 'PostmanRuntime/7.30.0'
    }    
    df1 = get_data_from_api(url1, payload1, headers)
    df2 = get_data_from_api(url2, payload2, headers)


    df1 = df1.add_prefix('People_')
    df2 = df2.add_prefix('Just_')


    combined_df = pd.concat([df1, df2], axis=1)


    selected_columns = [
    'People_Id', 'People_Matricula', 'People_Nome', 'People_Cpf',
    'Just_Id', 'Just_Description',
    ]
    
    output_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",  # extensão padrão
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],  # tipos de arquivo
        title="Salvar arquivo como"
    )

    # Cria um DataFrame final apenas com as colunas selecionadas
    final_df = combined_df[selected_columns]
    
    final_df.to_excel(output_path, index=False)

# Cria um Workbook do openpyxl
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    # Ajusta a largura das colunas automaticamente
    for col in ws.columns:
        max_length = 0
    column = col[0].column_letter
    for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
            except:
                pass
            
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    wb.save(output_path)
    exibir_log(f'Coleta de Justificativas salva em {output_path}')

def coleta_empresa():
    url1= 'https://www.mdcomune.com.br/RestServiceApi/CalculationRules/GetCalculationRulesSummary' # Cálculo
    url2= 'https://www.mdcomune.com.br/RestServiceApi/Schedules/GetSchedulesSummary' # Horário
    url3= 'https://www.mdcomune.com.br/RestServiceApi/OrganizationalStructure/GetOrganizationalStructure' # Estrutura
    url4= 'https://www.mdcomune.com.br/RestServiceApi/JobPosition/SearchJobPosition' # Cargo
    url5= 'https://www.mdcomune.com.br/RestServiceApi/Company/GetCompany' # Empresa
    
    payload1={}
    payload2 = {"ResponseType": "AS400V1"}
    
    headers = {
        'identifier': dados_selecionados["CNPJ"],
        'key': dados_selecionados['Chave API'],
        'User-Agent': 'PostmanRuntime/7.30.0'
    }
    
    df1 = get_data_from_api(url1, payload1, headers)
    df2 = get_data_from_api(url2, payload1, headers)
    df3 = get_data_from_api(url3, payload1, headers)  # Já tratado dentro da função
    df4 = get_data_from_api(url4, payload1, headers)
    df5 = get_data_from_api(url5, payload2, headers)
    
    # Adiciona prefixos para diferenciar as colunas dos DataFrames
    df1 = df1.add_prefix('Calculo_')
    df2 = df2.add_prefix('Horario_')
    df4 = df4.add_prefix('Cargo_')
    df5 = df5.add_prefix('Empresa_')

    # Combina os DataFrames horizontalmente (por colunas)
    combined_df = pd.concat([df1, df2, df3, df4, df5], axis=1)

    # Remove espaços extras dos nomes das colunas
    combined_df.columns = combined_df.columns.str.strip()

    # Seleciona as colunas desejadas, verificando se estão presentes
    selected_columns = [
        'Empresa_name', 'Empresa_CnpjCpf', 'Calculo_Id', 'Calculo_Descricao', 'Horario_Id', 'Horario_Descricao', 'Id', 'Description',
        'Cargo_Id', 'Cargo_Descricao'
    ]
    selected_columns = [col for col in selected_columns if col in combined_df.columns]
    
    for col in selected_columns:
        if col not in combined_df.columns:
            combined_df[col] = None
    
    
    output_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",  # extensão padrão
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],  # tipos de arquivo
        title="Salvar arquivo como"
    )

    # Cria um DataFrame final apenas com as colunas selecionadas
    final_df = combined_df[selected_columns]
    
    first_sheet_df = pd.DataFrame(columns=["Matricula", "Cracha", "Nome Completo", "CPF", "Admissão", "Nascimento", "Base de Horas",
                                        "Estrutura", "Horário", "Cálculo", "Cargo", "Possui PIS?", "PIS", "Sexo"])

    # Escrevendo múltiplas planilhas no mesmo arquivo Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        first_sheet_df.to_excel(writer, sheet_name='Cadastro', index=False)
        final_df.to_excel(writer, sheet_name='Dados Empresa', index=False)
        
    # Carrega o Workbook do openpyxl para adicionar validações e ajustar colunas
    wb = openpyxl.load_workbook(output_path)
    ws = wb['Cadastro']

    # Ajusta a largura das colunas automaticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Pega a letra da coluna
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="196B24", end_color="196B24", fill_type="solid")


    ws = wb['Dados Empresa']

    # Ajusta a largura das colunas automaticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Pega a letra da coluna
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2
        
    ws.protection.sheet = True  # Ativa a proteção da planilha
    ws.protection.password = "123456789a"  # Define uma senha (opcional)

    # Impede alterações de conteúdo, mas permite dimensionar células e selecionar
    ws.protection.enable()

    # Ajustes opcionais de permissões
    ws.protection.allow_format_cells = True  # Permite formatar células
    ws.protection.allow_select_locked_cells = True  # Permite selecionar células bloqueadas

    # Salva o arquivo Excel com as alterações
    wb.save(output_path)
    
    exibir_log(f'Planilha de Cadastro salva em {output_path}')

def alteracao_pessoas():
    url = 'https://www.mdcomune.com.br/RestServiceApi/People/SearchPeople'
    payload = {
        "Matricula": 0 
    }
    headers = {
        "identifier": dados_selecionados['CNPJ'],
        "key": dados_selecionados['Chave API'],
        'User-Agent': 'PostmanRuntime/7.30.0'
    }
    
    df = get_data_from_api(url, payload, headers)
    df = df.add_prefix('People_')
    
    selected_colomns =[
        'People_Id', 'People_Matricula', 'People_Cracha', 'People_Nome', 'People_DataNascimento', 'People_DataAdmissao',
    'People_Rg', 'People_Cpf', 'People_Email', 'People_BaseHoras', 'People_CodigoPis', 'People_Sexo', 'People_PessoaStatus', 'People_Estrutura.Id', 'People_Estrutura.Codigo',  'People_TipoSalario.Id',
    'People_Cargo.Id'
    ]
    
    # Verifica e preenche colunas ausentes com valores None
    for col in selected_colomns:
        if col not in df.columns:
            df[col] = None
    
    
    df_filtrado = df[selected_colomns]
    
    output_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",  # extensão padrão
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],  # tipos de arquivo
        title="Salvar arquivo como"
    )
    
    df_filtrado.to_excel(output_path, index=False)
    
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    for col in ws.columns:
        max_length = 0
    column = col[0].column_letter
    for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
            except:
                pass
            
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    wb.save(output_path)
    
    exibir_log(f'Planilha de funcionários salva em {output_path}')

def alteracao_pessoas_envio():
    
    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione o arquivo de cadastro de pessoas",
        filetypes=[("Arquivo Excel", "*.xlsx *.xls")]
    )
    if not caminho_arquivo:
        return

    try:
        df = pd.read_excel(caminho_arquivo)
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao ler o arquivo: {e}")
        return

    resultados = []
    df = df.fillna('')  # Preenche NaNs com string vazia
    url = 'https://www.mdcomune.com.br/RestServiceApi/People/ChangePerson'
    headers = {
        'identifier': dados_selecionados['CNPJ'],
        'key': dados_selecionados['Chave API'],
        'User-Agent': 'PostmanRuntime/7.30.0'
    }

    for index, row in df.iterrows():
        try:
            Id = row['People_Id']
            Matricula = row['People_Matricula']
            Cracha = row['People_Cracha']
            Nome = row['People_Nome']
            DataAdm = row['People_DataAdmissao']
            Rg = row.get('People_Rg', ' ')
            Cpf = row['People_Cpf'].replace(".","").replace("-","")
            
            # Verifica se a coluna Email existe no DataFrame
            if 'People_Email' in df.columns:
                Email = row['People_Email']
                Email = str(Email).strip() if pd.notna(Email) else None  # Converte para string e remove espaços
                if Email == '':
                    Email = None  # Define como None se a string for vazia
            else:
                Email = None  # Coluna não existe, então Email é None

            BaseHoras = row['People_BaseHoras']
            Pis = row['People_CodigoPis']
            Sexo = row['People_Sexo']
            EstruturaId = row['People_Estrutura.Id']
            TipoSalario = row['People_TipoSalario.Id']
            Cargo = row['People_Cargo.Id']

            # Verifica e converte o PIS para string
            Pis = None if pd.isna(Pis) else str(Pis)

            if Pis and len(Pis) == 11:
                flag_pis = True 
            elif Pis and len(Pis) == 12 and Pis.startswith('9'):
                flag_pis = False  
            else:
                flag_pis = False  

            # Monta o payload
            payload = {
                "Id": Id,
                "Matricula": Matricula,
                "Cracha": Cracha,
                "Nome": Nome,
                "DataAdmissao": DataAdm,
                "Rg": Rg,
                "Cpf": Cpf,
                "BaseHoras": BaseHoras,
                "Estrutura": {"Id": EstruturaId},
                "TipoFuncionario": {"Id": 1},
                "TipoSalario": {"Id": 101},
                "FlagGerarNumeroPISAutomatico": bool(flag_pis == False),
                "Sexo": Sexo,
            }

            # Exclui o Email do payload se for None
            if Email is not None:
                payload["Email"] = Email

            print(payload)  # Para verificar o payload antes de enviar
            response = requests.post(url, json=payload, headers=headers)
            response_json = response.json()

            # Verificar sucesso ou falha
            if response_json.get("Sucesso"):
                status = "Sucesso"
                mensagem = response_json.get("Mensagem", "")
            else:
                status = "Falha"
                mensagem = response_json.get("Mensagem", "Erro desconhecido")

        except Exception as e:
            status = "Erro"
            mensagem = str(e)

        # Armazenar o resultado
        resultados.append([Matricula, status, mensagem])
        
        exibir_log(f"Matricula: {Matricula} | Status: {status} | Mensagem: {mensagem}")
    
def exibir_log(mensagem):
    """Exibe o log de mensagens na interface."""
    log_widget.insert(tk.END, mensagem + "\n")
    log_widget.see(tk.END)

def confirmar_selecao():
    """Confirma a seleção da empresa e salva as credenciais globalmente."""
    global dados_selecionados
    if not razao_social_var.get():
        messagebox.showwarning("Aviso", "Selecione uma empresa.")
        return
    
    cpf_responsavel = cpf_var.get()[:11]
    cpf_var.set(cpf_responsavel)  # Atualiza a variável com o valor truncado
    
    dados_selecionados = {
        "Razão Social": razao_social_var.get(),
        "CNPJ": cnpj_var.get(),
        "Chave API": chave_var.get(),
        "CPF Responsável": cpf_responsavel,
        "Data Início" : data_inicio_var.get(),
        "Data fim": data_fim_var.get()
    }
    messagebox.showinfo("Sucesso", "Informações selecionadas com sucesso!")

def selecionar_arquivo_empresas():
    """Abre uma janela para o usuário selecionar o arquivo de empresas."""
    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione o arquivo Empresas",
        filetypes=[("Arquivos Excel", "*.xlsx")]
    )
    if not caminho_arquivo:
        raise FileNotFoundError("Nenhum arquivo selecionado.")
    return caminho_arquivo

def formatar_data(entry, new_value):
    # Remove tudo que não for número
    new_value = ''.join([c for c in new_value if c.isdigit()])
    
    # Limita a quantidade de caracteres para 8 (ddmmyyyy)
    if len(new_value) > 8:
        new_value = new_value[:8]

    # Aplica a formatação de data dd/mm/aaaa
    if len(new_value) > 2:
        new_value = new_value[:2] + '/' + new_value[2:]
    if len(new_value) > 5:
        new_value = new_value[:5] + '/' + new_value[5:]

    # Atualiza a entrada com a data formatada
    entry.delete(0, tk.END)
    entry.insert(0, new_value)

root = ttk.Window(themename="darkly")  # ttkbootstrap com tema
root.title("Seleção de Empresa e Envio de Marcações")
root.iconbitmap("M-Comune.ico")

# Carrega o arquivo de empresas
df_empresas = pd.read_excel(selecionar_arquivo_empresas())  # Ajuste com o caminho do seu Excel

razao_social_var = tk.StringVar()
cnpj_var = tk.StringVar()
chave_var = tk.StringVar()
cpf_var = tk.StringVar()
data_inicio_var = tk.StringVar()
data_fim_var = tk.StringVar()

frame_selecao = ttk.Frame(root, padding=10)
frame_selecao.grid(row=0, column=0, pady=10, padx=10, sticky="nsew")
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)

frame_campos = ttk.Frame(frame_selecao)
frame_campos.grid(row=0, column=0, pady=10, padx=10)


# Adicionando a Label e a Combobox para Razão Social
ttk.Label(frame_campos, text="Selecione a Razão Social:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
combo_razao_social = ttk.Combobox(frame_campos, textvariable=razao_social_var, width=40)
combo_razao_social['values'] = df_empresas["Razão Social"].tolist()
combo_razao_social.grid(row=0, column=1, padx=5, pady=5)
combo_razao_social.bind("<<ComboboxSelected>>", lambda e: preencher_detalhes())

# Adicionando a Label e a Entry para CNPJ
ttk.Label(frame_campos, text="CNPJ:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
ttk.Entry(frame_campos, textvariable=cnpj_var, state='readonly', width=42).grid(row=1, column=1, padx=5, pady=5)

# Adicionando a Label e a Entry para Chave API
ttk.Label(frame_campos, text="Chave API:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
ttk.Entry(frame_campos, textvariable=chave_var, state='readonly', width=42).grid(row=2, column=1, padx=5, pady=5)

# Adicionando a Label e a Entry para CPF Responsável
ttk.Label(frame_campos, text="CPF Responsável:").grid(row=3, column=0, padx=5, pady=5, sticky="w")
ttk.Entry(frame_campos, textvariable=cpf_var, state='readonly', width=42).grid(row=3, column=1, padx=5, pady=5)

# Interface de seleção de datas com formatação automática
ttk.Label(frame_campos, text="Data Início:").grid(row=4, column=0, padx=5, pady=5, sticky="w")
entry_data_inicio = ttk.Entry(frame_campos, textvariable=data_inicio_var, width=10)
entry_data_inicio.grid(row=4, column=1, padx=5, pady=5, sticky="w")
entry_data_inicio.bind("<KeyRelease>", lambda event: formatar_data(entry_data_inicio, entry_data_inicio.get()))

ttk.Label(frame_campos, text="Data Fim:").grid(row=5, column=0, padx=5, pady=5, sticky="w")
entry_data_fim = ttk.Entry(frame_campos, textvariable=data_fim_var, width=10)
entry_data_fim.grid(row=5, column=1, padx=5, pady=5, sticky="w")
entry_data_fim.bind("<KeyRelease>", lambda event: formatar_data(entry_data_fim, entry_data_fim.get()))

ttk.Button(frame_campos, text="Confirmar Seleção", command=confirmar_selecao, bootstyle="success").grid(row=5, column=2, padx=5, pady=5)

# Centralizando os botões
botao_frame = ttk.Frame(frame_selecao)
botao_frame.grid(row=6, column=0, columnspan=2, pady=15)
ttk.Button(botao_frame, text="Enviar Marcações", command=enviar_marcacoes, bootstyle="success").grid(row=0, column=1, padx=10)
ttk.Button(botao_frame, text="Envio", command=abrir_popup_selecao_pessoas, bootstyle="success").grid(row=0, column=2, padx=10)
ttk.Button(botao_frame, text="Coleta", command=abrir_popup_selecao_coleta, bootstyle="success").grid(row=0, column=3, padx=10)

# Ajustes para centralização geral do frame
frame_selecao.grid_columnconfigure(0, weight=1)
frame_selecao.grid_rowconfigure(0, weight=1)
frame_campos.grid_columnconfigure(0, weight=1)
frame_campos.grid_columnconfigure(1, weight=1)
frame_campos.grid_columnconfigure(2, weight=1)

# Log de mensagens centralizado e expandido
log_widget = tk.Text(frame_selecao, height=10, width=80)
log_widget.grid(row=7, column=0, columnspan=2, pady=15, padx=5, sticky="we")

root.mainloop()

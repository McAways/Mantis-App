import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import ttkbootstrap as ttk
from ttkbootstrap import DateEntry
from ttkbootstrap.constants import *
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import requests
import json
from datetime import datetime, timedelta
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import re
import threading
import os

dados_selecionados = {}

#Funções de Tratamento de ponto (Coleta)

def coleta_planilha_marcacoes_faltantes():
    output_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        title="Salvar planilha de faltas e atrasos em:"
    )

    if api_var.get() == "Kairos":
        url = 'https://www.dimepkairos.com.br/RestServiceApi/ReportEmployeePunch/GetReportEmployeePunch'
    else:
        url = 'https://www.mdcomune.com.br/RestServiceApi/ReportEmployeePunch/GetReportEmployeePunch'
    
    headers = {
        "identifier": dados_selecionados["CNPJ"],
        "key": dados_selecionados["Chave API"],
        'User-Agent': 'PostmanRuntime/7.30.0'
    }

    payload = {
        "MatriculaPessoa": [],
        "DataInicio": dados_selecionados["Data Início"],
        "DataFim": dados_selecionados["Data fim"],
        "ResponseType":"AS400V1"
    }
    
    try:
        # Verificar o intervalo de datas
        start_date = datetime.strptime(payload["DataInicio"], "%d/%m/%Y")
        end_date = datetime.strptime(payload["DataFim"], "%d/%m/%Y")
        delta_days = (end_date - start_date).days
        
        if delta_days > 90:
            messagebox.showerror(
                "Erro de Intervalo de Datas",
                "O intervalo de datas selecionado é maior que 90 dias. Por favor, selecione um período menor."
            )
            exibir_log("Erro: Intervalo de datas maior que 90 dias.")
            return

    except Exception as e:
        messagebox.showerror(
            "Erro de Datas",
            f"Erro ao validar as datas: {e}"
        )
        return
    
    response = requests.post(url, json=payload, headers=headers)

    if response.status_code == 200:
        try:
            data = response.json()
            if "Obj" in data and isinstance(data['Obj'], list):
                all_entries = []
                all_fixed_data = []

                start_date = datetime.strptime(payload["DataInicio"], "%d/%m/%Y")
                end_date = datetime.strptime(payload["DataFim"], "%d/%m/%Y")
                full_date_range = generate_date_range(start_date, end_date)

                justificativas = carregar_justificativas_salvas(dados_selecionados["CNPJ"])

                for item in data['Obj']:
                    fixed_data = {
                        'Funcionario': item['InfoFuncionario']['Nome'],
                        'PIS': item['InfoFuncionario']['PIS'],
                        'Matricula': item['InfoFuncionario']['Matricula'],
                        'Estrutura': item['InfoFuncionario']['Estrutura']
                    }

                    entradas_por_data = {}
                    for entrada in item['Entradas']:
                        data_limpa = clean_json_date(entrada['Data'])
                        entrada_data_formatada = datetime.strptime(data_limpa, "%d/%m/%Y")
                        justificativa_existente = entrada.get('Justificativa', '').strip()
                        entradas_por_data[entrada_data_formatada] = entrada
                    for date in full_date_range:
                        if date in entradas_por_data:
                            entrada = entradas_por_data[date]
                            entry_data = {
                                'Data': date.strftime("%d/%m/%Y"),
                                'Horario': entrada['Horario'],
                                'Apontamentos': entrada['Apontamentos'],
                                'HTrab': entrada['HTrab'],
                                'Descontos': entrada['Descontos'],
                                'Debito': entrada['Debito'],
                                'Justificativa': justificativa_existente
                            }
                        else:
                            entry_data = {
                                'Data': date.strftime("%d/%m/%Y"),
                                'Horario': '',
                                'Apontamentos': ''
                            }

                        combined_data = {**fixed_data, **entry_data}
                        all_entries.append(combined_data)

                final_df = pd.DataFrame(all_entries)
                faltas = process_faltas(final_df)
                
                if not faltas.empty:
                    print("Faltas identificadas:")
                    print(faltas)
                else:
                    print("Nenhuma falta identificada.")

                final_df = faltas
                final_df['Qtd Horas'] = ''
                final_df['Entrada'] = ''
                final_df['Almoço Ida'] = ''
                final_df['Almoço Volta'] = ''
                final_df['Saida'] = ''
                final_df['Empresa'] = item['InfoEmpresa']['Nome']

                final_df.to_excel(output_path, index=False)

                wb = openpyxl.load_workbook(output_path)
                ws = wb.active
                
                justificativa_str = 'Folga,' + ','.join(justificativas)
                justificativa_validation = DataValidation(
                    type='list',
                    formula1=f'"{justificativa_str}"',
                    allow_blank=True
                )
                justificativa_validation.error = 'Escolher valores da lista'
                justificativa_validation.errorTitle = 'Entrada Invalida'
                justificativa_validation.prompt = 'Selecione uma justificativa'
                justificativa_validation.promptTitle = 'Justificativas'
                justificativa_col_index = final_df.columns.get_loc("Justificativa") + 1
                justificativa_col_letter = get_column_letter(justificativa_col_index)
                ws.add_data_validation(justificativa_validation)
                justificativa_validation.add(f'{justificativa_col_letter}2:{justificativa_col_letter}{len(final_df)+1}')

                for col in ws.columns:
                                max_length = 0
                                column = col[0].column_letter
                                for cell in col:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(cell.value)
                                    except:
                                        pass
                                adjusted_width = (max_length + 2)
                                ws.column_dimensions[column].width = adjusted_width

                fill_grey = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                            
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        if cell.row % 2 == 0:
                            cell.fill = fill_grey
                        else:
                            cell.fill = fill_white
                            cell.alignment = Alignment(horizontal="left", vertical="center")

                apply_borders(ws)
                wb.save(output_path)
                exibir_log(f"Arquivo Excel salvo com sucesso em {output_path}")
            else:
                exibir_log("Nenhum dado no campo 'Obj'.")
        except ValueError as e:
            exibir_log(f"Erro ao decodificar JSON: {e}")
    else:
        exibir_log(f"Falha: {response.status_code} - {response.text}")

def coleta_planilha_marcacoes_inconsistencia():
    output_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                    title="Salvar planilha de inconsistências em:"
                    )
    
    if api_var.get() == "Kairos":
        url = 'https://www.dimepkairos.com.br/RestServiceApi/ReportEmployeePunch/GetReportEmployeePunch'
    else:
        url = 'https://www.mdcomune.com.br/RestServiceApi/ReportEmployeePunch/GetReportEmployeePunch'
    
    headers = {
            "identifier": dados_selecionados["CNPJ"],
            "key": dados_selecionados["Chave API"],
            'User-Agent': 'PostmanRuntime/7.30.0'
    }

    payload = {
            "MatriculaPessoa": [],
            "DataInicio": dados_selecionados["Data Início"],
            "DataFim": dados_selecionados["Data fim"],
            "ResponseType": "AS400V1"
    }
    
    try:
        # Verificar o intervalo de datas
        start_date = datetime.strptime(payload["DataInicio"], "%d/%m/%Y")
        end_date = datetime.strptime(payload["DataFim"], "%d/%m/%Y")
        delta_days = (end_date - start_date).days
        
        if delta_days > 90:
            messagebox.showerror(
                "Erro de Intervalo de Datas",
                "O intervalo de datas selecionado é maior que 90 dias. Por favor, selecione um período menor."
            )
            exibir_log("Erro: Intervalo de datas maior que 90 dias.")
            return

    except Exception as e:
        messagebox.showerror(
            "Erro de Datas",
            f"Erro ao validar as datas: {e}"
        )
        return
    
    response = requests.post(url, json=payload, headers=headers)

    if response.status_code == 200:
        try:
            data = response.json()
            if "Obj" in data and isinstance(data['Obj'], list) and len(data['Obj']) > 0:
                all_entries = []
                
                justificativas = carregar_justificativas_salvas(dados_selecionados["CNPJ"])
                
                for item in data['Obj']:
                    fixed_data = {
                        'Funcionario': item['InfoFuncionario']['Nome'],
                        'Matricula': item['InfoFuncionario']['Matricula'],
                        'Estrutura': item['InfoFuncionario']['Estrutura']
                    }

                    for entrada in item['Entradas']:
                        data_limpa = clean_json_date(entrada['Data'])
                        entrada_data_formatada = datetime.strptime(data_limpa, "%d/%m/%Y")
                        justificativa_existente = entrada.get('Justificativa', '').strip()
                        marcacoes_impares = filtra_marcacoes_impares_e_htrab_vazio([entrada])
                        if marcacoes_impares:
                            for marcacao in marcacoes_impares:
                                entry_data = {
                                    'Data': entrada_data_formatada.strftime("%d/%m/%Y"),
                                    'Horario': marcacao['Horario'],
                                    'Apontamentos': marcacao['Apontamentos'],
                                    'HTrab': entrada['HTrab'],
                                    'Desconto': entrada['Descontos'],
                                    'Debito': entrada['Debito'],
                                    'Justificativa': justificativa_existente
                                    
                                }
                                combined_data = {**fixed_data, **entry_data}
                                all_entries.append(combined_data)

                if all_entries:
                    final_df = pd.DataFrame(all_entries)

                    final_df['Qtd Horas'] = ''
                    final_df['Entrada'] = ''
                    final_df['Almoço Ida'] = ''
                    final_df['Almoço Volta'] = ''
                    final_df['Saida'] = ''
                    final_df['Empresa'] = item['InfoEmpresa']['Nome']
                    
                    inconsistencia_df = final_df
                    
                    inconsistencia_df.to_excel(output_path, index=False)

                wb = openpyxl.load_workbook(output_path)
                ws = wb.active

                justificativa_str = 'Folga,' + ','.join(justificativas)
                justificativa_validation = DataValidation(
                    type='list',
                    formula1=f'"{justificativa_str}"',
                    allow_blank=True
                )
                justificativa_validation.error = 'Escolher valores da lista'
                justificativa_validation.errorTitle = 'Entrada Invalida'
                justificativa_validation.prompt = 'Selecione uma justificativa'
                justificativa_validation.promptTitle = 'Justificativas'
                justificativa_col_index = final_df.columns.get_loc("Justificativa") + 1
                justificativa_col_letter = get_column_letter(justificativa_col_index)
                ws.add_data_validation(justificativa_validation)
                justificativa_validation.add(f'{justificativa_col_letter}2:{justificativa_col_letter}{len(final_df)+1}')

                for col in ws.columns:
                                max_length = 0
                                column = col[0].column_letter
                                for cell in col:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(cell.value)
                                    except:
                                        pass
                                adjusted_width = (max_length + 2)
                                ws.column_dimensions[column].width = adjusted_width

                fill_grey = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                            
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        if cell.row % 2 == 0:
                            cell.fill = fill_grey
                        else:
                            cell.fill = fill_white
                            cell.alignment = Alignment(horizontal="left", vertical="center")

                apply_borders(ws)
                wb.save(output_path)
                exibir_log(f"Arquivo Excel salvo com sucesso em {output_path}")
            else:
                exibir_log("Nenhum dado no campo 'Obj'.")
        except ValueError as e:
            exibir_log(f"Erro ao decodificar JSON: {e}")
    else:
        exibir_log(f"Falha: {response.status_code} - {response.text}")

def coleta_planilha_marcacoes_incomum():
    output_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                    title="Salvar planilha de marcação incomum em:"
                    )
    
    if api_var.get() == "Kairos":
        url = 'https://www.dimepkairos.com.br/RestServiceApi/ReportEmployeePunch/GetReportEmployeePunch'
    else:
        url = 'https://www.mdcomune.com.br/RestServiceApi/ReportEmployeePunch/GetReportEmployeePunch'
    
    headers = {
            "identifier": dados_selecionados["CNPJ"],
            "key": dados_selecionados["Chave API"],
            'User-Agent': 'PostmanRuntime/7.30.0'
    }

    payload = {
            "MatriculaPessoa": [],
            "DataInicio": dados_selecionados["Data Início"],
            "DataFim": dados_selecionados["Data fim"],
            "ResponseType": "AS400V1"
    }
    
    try:
        # Verificar o intervalo de datas
        start_date = datetime.strptime(payload["DataInicio"], "%d/%m/%Y")
        end_date = datetime.strptime(payload["DataFim"], "%d/%m/%Y")
        delta_days = (end_date - start_date).days
        
        if delta_days > 90:
            messagebox.showerror(
                "Erro de Intervalo de Datas",
                "O intervalo de datas selecionado é maior que 90 dias. Por favor, selecione um período menor."
            )
            exibir_log("Erro: Intervalo de datas maior que 90 dias.")
            return

    except Exception as e:
        messagebox.showerror(
            "Erro de Datas",
            f"Erro ao validar as datas: {e}"
        )
        return
    
    response = requests.post(url, json=payload, headers=headers)

    if response.status_code == 200:
        try:
            data = response.json()
            if "Obj" in data and isinstance(data['Obj'], list) and len(data['Obj']) > 0:
                all_entries = []
                
                justificativas = carregar_justificativas_salvas(dados_selecionados["CNPJ"])
                
                for item in data['Obj']:
                    fixed_data = {
                        'Funcionario': item['InfoFuncionario']['Nome'],
                        'Matricula': item['InfoFuncionario']['Matricula'],
                        'Estrutura': item['InfoFuncionario']['Estrutura']
                    }
                    

                    for entrada in item['Entradas']:
                        data_limpa = clean_json_date(entrada['Data'])
                        entrada_data_formatada = datetime.strptime(data_limpa, "%d/%m/%Y")
                        justificativa_existente = entrada.get('Justificativa', '').strip()

                        entry_data = {
                            'Data': entrada_data_formatada.strftime("%d/%m/%Y"),
                            'Horario': entrada['Horario'],
                            'Apontamentos': entrada['Apontamentos'],
                            'HTrab': entrada['HTrab'],
                            'Desconto': entrada['Descontos'],
                            'Debito': entrada['Debito'],
                            'Justificativa': justificativa_existente
                        }
                        combined_data = {**fixed_data, **entry_data}
                        all_entries.append(combined_data)

                if all_entries:
                    final_df = pd.DataFrame(all_entries)
                    final_df['Qtd Horas'] = ''
                    final_df['Entrada'] = ''
                    final_df['Almoço Ida'] = ''
                    final_df['Almoço Volta'] = ''
                    final_df['Saida'] = ''
                    final_df['Empresa']= item['InfoEmpresa']['Nome']

                    incomuns_df = process_incomum(final_df)
                    

                    if not incomuns_df.empty:
                        incomuns_df.to_excel(output_path, index=False)

                        wb = openpyxl.load_workbook(output_path)
                        ws = wb.active

                        justificativa_str = 'Folga,' + ','.join(justificativas)
                        justificativa_validation = DataValidation(
                                type='list',
                                formula1=f'"{justificativa_str}"',
                                allow_blank=True
                            )
                        justificativa_validation.error = 'Escolher valores da lista'
                        justificativa_validation.errorTitle = 'Entrada Invalida'
                        justificativa_validation.prompt = 'Selecione uma justificativa'
                        justificativa_validation.promptTitle = 'Justificativas'
                        justificativa_col_index = final_df.columns.get_loc("Justificativa") + 1
                        justificativa_col_letter = get_column_letter(justificativa_col_index)
                        ws.add_data_validation(justificativa_validation)
                        justificativa_validation.add(f'{justificativa_col_letter}2:{justificativa_col_letter}{len(final_df)+1}')
                            
                        for col in ws.columns:
                                max_length = 0
                                column = col[0].column_letter
                                for cell in col:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(cell.value)
                                    except:
                                        pass
                                adjusted_width = (max_length + 2)
                                ws.column_dimensions[column].width = adjusted_width

                        fill_grey = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                        fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                            
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                            for cell in row:
                                if cell.row % 2 == 0:
                                    cell.fill = fill_grey
                                else:
                                    cell.fill = fill_white
                                    cell.alignment = Alignment(horizontal="left", vertical="center")

                apply_borders(ws)
                wb.save(output_path)
                exibir_log(f"Arquivo Excel salvo com sucesso em {output_path}")
            else:
                exibir_log("Nenhum dado no campo 'Obj'.")
        except ValueError as e:
            exibir_log(f"Erro ao decodificar JSON: {e}")
    else:
        exibir_log(f"Falha: {response.status_code} - {response.text}")

def coleta_planilha_marcacoes():
    output_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        title="Salvar planilha de marcação em:"
    )
    if api_var.get() == "Kairos":
        url = 'https://www.dimepkairos.com.br/RestServiceApi/ReportEmployeePunch/GetReportEmployeePunch'
    else:
        url = 'https://www.mdcomune.com.br/RestServiceApi/ReportEmployeePunch/GetReportEmployeePunch'
    headers = {
        "identifier": dados_selecionados["CNPJ"],
        "key": dados_selecionados["Chave API"],
        'User-Agent': 'PostmanRuntime/7.30.0'
    }

    payload = {
        "MatriculaPessoa": [],
        "DataInicio": dados_selecionados["Data Início"],
        "DataFim": dados_selecionados["Data fim"],
        "ResponseType": "AS400V1"
    }

    try:
        start_date = datetime.strptime(payload["DataInicio"], "%d/%m/%Y")
        end_date = datetime.strptime(payload["DataFim"], "%d/%m/%Y")
        delta_days = (end_date - start_date).days

        if delta_days > 90:
            messagebox.showerror(
                "Erro de Intervalo de Datas",
                "O intervalo de datas selecionado é maior que 90 dias. Por favor, selecione um período menor."
            )
            exibir_log("Erro: Intervalo de datas maior que 90 dias.")
            return

    except Exception as e:
        messagebox.showerror("Erro de Datas", f"Erro ao validar as datas: {e}")
        return

    response = requests.post(url, json=payload, headers=headers)

    if response.status_code == 200:
        try:
            data = response.json()
            if "Obj" in data and isinstance(data['Obj'], list):
                all_entries = []
                full_date_range = generate_date_range(start_date, end_date)

                # Ajuste aqui para carregar justificativas com o nome correto
                justificativas = carregar_justificativas_salvas(dados_selecionados["CNPJ"])

                if not justificativas:
                    messagebox.showwarning("Aviso", "Nenhuma justificativa salva encontrada. Por favor, selecione justificativas.")
                    exibir_log("Nenhuma justificativa encontrada para a planilha.")
                    return

                for item in data['Obj']:
                    fixed_data = {
                        'Funcionario': item['InfoFuncionario']['Nome'],
                        'PIS': item['InfoFuncionario']['PIS'],
                        'Matricula': item['InfoFuncionario']['Matricula'],
                        'Estrutura': item['InfoFuncionario']['Estrutura']
                    }
                    entradas_por_data = {}
                    for entrada in item['Entradas']:
                        data_limpa = clean_json_date(entrada['Data'])
                        entrada_data_formatada = datetime.strptime(data_limpa, "%d/%m/%Y")
                        entradas_por_data[entrada_data_formatada] = entrada
                    for date in full_date_range:
                        justificativa_existente = entrada.get('Justificativa', '').strip()
                        if date in entradas_por_data:
                            entrada = entradas_por_data[date]
                            entry_data = {
                                'Data': date.strftime("%d/%m/%Y"),
                                'Horario': entrada['Horario'],
                                'Apontamentos': entrada['Apontamentos'],
                                'HTrab': entrada['HTrab'],
                                'Desconto': entrada['Descontos'],
                                'Debito': entrada['Debito'],
                                'Justificativa': justificativa_existente
                            }
                        else:
                            entry_data = {
                                'Data': date.strftime("%d/%m/%Y"),
                                'Horario': '',
                                'Apontamentos': '',
                                'Justificativa': ''
                            }

                        combined_data = {**fixed_data, **entry_data}
                        all_entries.append(combined_data)

                final_df = pd.DataFrame(all_entries)
                final_df['Qtd Horas'] = ''
                final_df['Entrada'] = ''
                final_df['Almoço Ida'] = ''
                final_df['Almoço Volta'] = ''
                final_df['Saida'] = ''
                final_df['Empresa'] = item['InfoEmpresa']['Nome']
                final_df.to_excel(output_path, index=False)

                wb = openpyxl.load_workbook(output_path)
                ws = wb.active

                justificativa_str = 'Folga,' + ','.join(justificativas)
                justificativa_validation = DataValidation(
                    type='list',
                    formula1=f'"{justificativa_str}"',
                    allow_blank=True
                )
                justificativa_validation.error = 'Escolher valores da lista'
                justificativa_validation.errorTitle = 'Entrada Invalida'
                justificativa_validation.prompt = 'Selecione uma justificativa'
                justificativa_validation.promptTitle = 'Justificativas'
                justificativa_col_index = final_df.columns.get_loc("Justificativa") + 1
                justificativa_col_letter = get_column_letter(justificativa_col_index)
                ws.add_data_validation(justificativa_validation)
                justificativa_validation.add(f'{justificativa_col_letter}2:{justificativa_col_letter}{len(final_df)+1}')

                for col in ws.columns:
                                max_length = 0
                                column = col[0].column_letter
                                for cell in col:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(cell.value)
                                    except:
                                        pass
                                adjusted_width = (max_length + 2)
                                ws.column_dimensions[column].width = adjusted_width

                fill_grey = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                            
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        if cell.row % 2 == 0:
                            cell.fill = fill_grey
                        else:
                            cell.fill = fill_white
                            cell.alignment = Alignment(horizontal="left", vertical="center")

                apply_borders(ws)
                wb.save(output_path)
                exibir_log(f"Arquivo Excel salvo com sucesso em {output_path}")
            else:
                exibir_log("Nenhum dado no campo 'Obj'.")
        except ValueError as e:
            exibir_log(f"Erro ao decodificar JSON: {e}")
    else:
        exibir_log(f"Falha: {response.status_code} - {response.text}")

def coleta_empresa():
    if api_var.get() == "Kairos":
        url1= 'https://www.dimepkairos.com.br/RestServiceApi/CalculationRules/GetCalculationRulesSummary' # Cálculo
        url2= 'https://www.dimepkairos.com.br/RestServiceApi/Schedules/GetSchedulesSummary' # Horário
        url3= 'https://www.dimepkairos.com.br/RestServiceApi/OrganizationalStructure/GetOrganizationalStructure' # Estrutura
        url4= 'https://www.dimepkairos.com.br/RestServiceApi/JobPosition/SearchJobPosition' # Cargo
        url5= 'https://www.dimepkairos.com.br/RestServiceApi/Company/GetCompany' # Empresa
    else:
        url1= 'https://www.mdcomune.com.br/RestServiceApi/CalculationRules/GetCalculationRulesSummary' # Cálculo
        url2= 'https://www.mdcomune.com.br/RestServiceApi/Schedules/GetSchedulesSummary' # Horário
        url3= 'https://www.mdcomune.com.br/RestServiceApi/OrganizationalStructure/GetOrganizationalStructure' # Estrutura
        url4= 'https://www.mdcomune.com.br/RestServiceApi/JobPosition/SearchJobPosition' # Cargo
        url5= 'https://www.mdcomune.com.br/RestServiceApi/Company/GetCompany' # Empresa
        
    payload1={}
    payload2 = {"ResponseType": "AS400V1"}
    
    headers = {
        'identifier': dados_selecionados["CNPJ"],
        'key': dados_selecionados['Chave API'],
        'User-Agent': 'PostmanRuntime/7.30.0'
    }
    
    df1 = get_data_from_api(url1, payload1, headers)
    df2 = get_data_from_api(url2, payload1, headers)
    df3 = get_data_from_api(url3, payload1, headers)
    df4 = get_data_from_api(url4, payload1, headers)
    df5 = get_data_from_api(url5, payload2, headers)

    df1 = df1.add_prefix('Calculo_')
    df2 = df2.add_prefix('Horario_')
    df4 = df4.add_prefix('Cargo_')
    df5 = df5.add_prefix('Empresa_')
    
    combined_df = pd.concat([df1, df2, df3, df4, df5], axis=1)

    combined_df.columns = combined_df.columns.str.strip()

    selected_columns = [
        'Empresa_name', 'Empresa_CnpjCpf', 'Calculo_Id', 'Calculo_Descricao', 'Horario_Id', 'Horario_Descricao', 'Id', 'Description',
        'Cargo_Id', 'Cargo_Descricao'
    ]
    selected_columns = [col for col in selected_columns if col in combined_df.columns]
    
    for col in selected_columns:
        if col not in combined_df.columns:
            combined_df[col] = None
    
    
    output_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        title="Salvar arquivo como"
    )
    
    final_df = combined_df[selected_columns]
    
    first_sheet_df = pd.DataFrame(columns=["Matricula", "Cracha", "Nome Completo", "CPF", "RG","Email", "Admissão", "Nascimento", "Base de Horas",
        "Estrutura", "Horário", "Cálculo", "Cargo", "Possui PIS?", "PIS", "Sexo", "Campo Alternativo", "Endereço", "Número", "Bairro", "Cidade", "Estado", "Pais"])

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        first_sheet_df.to_excel(writer, sheet_name='Cadastro', index=False)
        final_df.to_excel(writer, sheet_name='Dados Empresa', index=False)
        
    wb = openpyxl.load_workbook(output_path)
    ws = wb['Cadastro']

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="196B24", end_color="196B24", fill_type="solid")


    ws = wb['Dados Empresa']

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2
        
    ws.protection.sheet = True
    ws.protection.password = "123456789a"

    ws.protection.enable()

    ws.protection.allow_format_cells = True
    ws.protection.allow_select_locked_cells = True

    wb.save(output_path)
    
    exibir_log(f'Planilha de Cadastro salva em {output_path}')

def coleta_planilha_desligamento():
    
    output_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        title="Salvar planilha de ferias em:"
    )
    if api_var.get() == "Kairos":
        url = 'https://www.dimepkairos.com.br/RestServiceApi/People/SearchPeople'
    else:
        url = 'https://www.mdcomune.com.br/RestServiceApi/People/SearchPeople'
        
    payload = {
        "Matricula": 0 
    }
    headers = {
        "identifier": dados_selecionados['CNPJ'],
        "key": dados_selecionados['Chave API'],
        'User-Agent': 'PostmanRuntime/7.30.0'
    }
    
    # Obtém os dados da API
    df = get_data_from_api(url, payload, headers)
    
    # Seleciona as colunas necessárias
    selected_columns = [
        'Matricula',
        'Nome',
        'DataDemissao'
    ]
    
    df_filtrado = df[selected_columns]
    
    # Filtra funcionários desligados
    df_filtrado = df_filtrado[df_filtrado['DataDemissao'] == "01/01/1753 00:00:00"]
    
    # Remove a coluna 'Colaborador_DataDemissao' para a planilha final
    df_filtrado = df_filtrado.drop(columns=['DataDemissao'])
    
    # Adiciona colunas de período de férias
    final_df = pd.DataFrame(df_filtrado)
    final_df['Data de Desligamento'] = ''
    
    # Salva o DataFrame filtrado em Excel
    final_df.to_excel(output_path, index=False)
    aplicar_estilo(output_path)
    
    exibir_log(f'Planilha de desligamento salvo em {output_path}')

def coleta_planilha_ferias():
    output_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        title="Salvar planilha de ferias em:"
    )
    
    if api_var.get() == "Kairos":
        url = 'https://www.dimepkairos.com.br/RestServiceApi/People/SearchPeople'
    else:
        url = 'https://www.mdcomune.com.br/RestServiceApi/People/SearchPeople'
    payload = {
        "Matricula": 0 
    }
    headers = {
        "identifier": dados_selecionados['CNPJ'],
        "key": dados_selecionados['Chave API'],
        'User-Agent': 'PostmanRuntime/7.30.0'
    }
    
    # Obtém os dados da API
    df = get_data_from_api(url, payload, headers)
    
    # Seleciona as colunas necessárias
    selected_columns = [
        'Matricula',
        'Nome',
        'DataDemissao'
    ]
    
    df_filtrado = df[selected_columns]
    
    # Filtra funcionários desligados
    df_filtrado = df_filtrado[df_filtrado['DataDemissao'] == "01/01/1753 00:00:00"]
    
    # Remove a coluna 'Colaborador_DataDemissao' para a planilha final
    df_filtrado = df_filtrado.drop(columns=['DataDemissao'])
    
    # Adiciona colunas de período de férias
    final_df = pd.DataFrame(df_filtrado)
    final_df['Data Inicio'] = ''
    final_df['Data Fim'] = ''
    
    # Salva o DataFrame filtrado em Excel
    final_df.to_excel(output_path, index=False)
    aplicar_estilo(output_path)
    
    exibir_log(f'Planilha de férias salva em {output_path}')

def alteracao_pessoas():
    """Exporta os dados de pessoas para um arquivo Excel para alterações."""
    output_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        title="Salvar Pessoas para alteração em:"
    )
    if not output_path:
        exibir_log("Nenhum local de salvamento foi selecionado.")
        return
    if api_var.get() == "Kairos":
        url = 'https://www.dimepkairos.com.br/RestServiceApi/People/SearchPeople'
    else:
        url = 'https://www.mdcomune.com.br/RestServiceApi/People/SearchPeople'
    payload = {"Matricula": 0}
    headers = {
        "identifier": dados_selecionados["CNPJ"],
        "key": dados_selecionados["Chave API"],
        "User-Agent": "PostmanRuntime/7.30.0"
    }

    try:
        # Obtém os dados da API
        df = get_data_from_api(url, payload, headers)
        if df.empty:
            exibir_log("Nenhum dado retornado pela API.")
            messagebox.showwarning("Atenção", "Nenhum dado foi retornado pela API.")
            return

        # Adiciona prefixo às colunas
        df = df.add_prefix('People_')

        # Colunas necessárias para o arquivo
        selected_columns = [
            'People_Id', 'People_Matricula', 'People_Cracha', 'People_Nome',
            'People_DataNascimento', 'People_DataAdmissao', 'People_Rg', 
            'People_Cpf', 'People_Email', 'People_BaseHoras', 'People_CodigoPis', 
            'People_Sexo', 'People_PessoaStatus', 'People_Estrutura.Id', 
            'People_Estrutura.Codigo', 'People_TipoSalario.Id', 'People_Cargo.Id'
        ]

        # Adiciona colunas ausentes ao DataFrame
        for col in selected_columns:
            if col not in df.columns:
                df[col] = None

        # Filtra as colunas desejadas
        df_filtrado = df[selected_columns]

        # Salva o DataFrame em Excel
        df_filtrado.to_excel(output_path, index=False)

        aplicar_estilo(output_path)
        exibir_log(f"Planilha de funcionários salva em {output_path}")
        messagebox.showinfo("Sucesso", f"Planilha salva em:\n{output_path}")

    except Exception as e:
        exibir_log(f"Erro ao processar: {e}")
        messagebox.showerror("Erro", f"Erro ao processar os dados: {e}")

def processar_marcacoes(df, barra, label_status, cancelar):
    if df.empty:
        exibir_log("Arquivo sem dados de marcações. Ignorando operação.")
        return
    if api_var.get() == "Kairos":
        url = "https://www.dimepkairos.com.br/RestServiceApi/Mark/SetMarks"
    else:
        url = "https://www.mdcomune.com.br/RestServiceApi/Mark/SetMarks"
    headers = {
        "identifier": dados_selecionados['CNPJ'],
        "key": dados_selecionados['Chave API'],
        "cpf": dados_selecionados['CPF Responsável'],
        'User-Agent': 'PostmanRuntime/7.30.0'
    }

    for index, row in df.iterrows():
        if cancelar.is_set():
            exibir_log("Envio de marcações cancelado pelo usuário.")
            return

        matricula = row.get('Matricula')
        data_completa = row.get('Data')
        horario_intervalo = row.get('Horario')  # Novo campo com o intervalo de horários

        marcacoes = {
            "Entrada": row.get('Entrada'),
            "Almoço Ida": row.get('Almoço Ida'),
            "Almoço Volta": row.get('Almoço Volta'),
            "Saida": row.get('Saida')
        }

        for tipo, hora in marcacoes.items():
            if cancelar.is_set():
                exibir_log("Envio de marcações cancelado pelo usuário.")
                return

            if pd.notna(hora):
                # Passa o intervalo de horário para a função combinar_data_hora
                data_hora_marcacao = combinar_data_hora(data_completa, hora, tipo, horario_intervalo)
                if data_hora_marcacao:
                    payload = {
                        "Matricula": matricula,
                        "DataHoraApontamento": data_hora_marcacao,
                        "CpfResponsavel": dados_selecionados['CPF Responsável'],
                        "ResponseType": "AS400V1"
                    }
                    response = requests.post(url, json=payload, headers=headers)
                    status = "Sucesso" if response.status_code == 200 else "Falha"
                    mensagem = response.json().get("Mensagem", "Erro desconhecido") if response.ok else response.text
                    exibir_log(f"Marcação de {tipo} enviada para matrícula {matricula} na data {data_hora_marcacao}: {status} - {mensagem}")

        if barra and not cancelar.is_set():
            barra.step(1)
            barra.update_idletasks()
        if label_status and not cancelar.is_set():
            label_status.config(text=f"Marcação enviada para Matricula: {matricula}")

    if not cancelar.is_set():
        exibir_log("Processamento de marcações concluído.")
        messagebox.showinfo("Sucesso", "Envio de marcações concluído.")
    else:
        exibir_log("Envio de marcações interrompido.")

def enviar_dados_combinados():
    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione a planilha de tratamento",
        filetypes=[("Arquivo Excel", "*.xlsx *.xls")]
    )
    if not caminho_arquivo:
        exibir_log(f"Nenhum arquivo selecionado. Operação de tratamento a {dados_selecionados['Razão Social']} abortada.")
        return

    try:
        df = pd.read_excel(caminho_arquivo)
    except Exception as e:
        exibir_log(f"Erro ao carregar o arquivo: {e}")
        return

    exibir_log(f"Iniciando o tratamento para {dados_selecionados['Razão Social']}")
    mostrar_duas_barras_progresso_paralelo(processar_marcacoes, processar_arquivo_excel, df, titulo="Enviando Dados")
    exibir_log(f"Tratamento da empresa {dados_selecionados['Razão Social']} concluído")

def coleta_cargos():
    try:
        output_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Salvar planilha de cargos em:"
        )
        
        if not output_path:
            exibir_log("Operação cancelada pelo usuário.")
            return
        if api_var.get() == "Kairos":
            url = 'https://www.dimepkairos.com.br/RestServiceApi/JobPosition/SearchJobPosition'
        else:
            url = 'https://www.mdcomune.com.br/RestServiceApi/JobPosition/SearchJobPosition'
        payload = {}
        headers = {
            "identifier": dados_selecionados['CNPJ'],
            "key": dados_selecionados['Chave API'],
            'User-Agent': 'PostmanRuntime/7.30.0'
        }
        
        df = get_data_from_api(url, payload, headers)
        
        if df.empty:
            exibir_log("Nenhum dado retornado pela API.")
            messagebox.showinfo("Aviso", "Nenhum cargo foi retornado pela API.")
            return

        required_columns = ['Id', 'Codigo', 'Descricao']
        selected_columns = [col for col in required_columns if col in df.columns]

        for col in required_columns:
            if col not in df.columns:
                df[col] = None

        final_df = df[required_columns]
        final_df_2 = pd.DataFrame(final_df)
        final_df_2["Codigo(Novo Cargo)"] = ''
        final_df_2["Descricao(Novo Cargo)"] = ''
        final_df_2.to_excel(output_path, index=False)
        aplicar_estilo(output_path)

        exibir_log(f'Planilha de cargos salva em {output_path}')
    
    except requests.RequestException as e:
        exibir_log(f"Erro ao conectar à API: {e}")
        messagebox.showerror("Erro", f"Erro ao conectar à API: {e}")
    
    except Exception as e:
        exibir_log(f"Erro durante a coleta de cargos: {e}")
        messagebox.showerror("Erro", f"Erro durante a coleta de cargos: {e}")
    
#Funções de Tratamento de ponto (Envio)

def enviar_justificativa(matricula, id_funcionario, id_justificativa, descricao, data, qtd_horas):
        if api_var.get() == "Kairos":
            url = "https://www.dimepkairos.com.br/RestServiceApi/PreJustificationRequest/PreJustificationRequest"
        else:
            url = "https://www.mdcomune.com.br/RestServiceApi/PreJustificationRequest/PreJustificationRequest"
        headers = {
            'identifier': dados_selecionados["CNPJ"],
            'key': dados_selecionados["Chave API"],
            'User-Agent': 'PostmanRuntime/7.30.0'
        }
        payload = {
            "IdJustification": id_justificativa,
            "IdUser": '1',
            "IdEmployee": id_funcionario,
            "QtdHours": qtd_horas,
            "Date": data,
            "Notes" : "Enviado via API",
            "RequestType":"1",
            "ResponseType":"AS400V1"
        }
        
        try:
            data_formatada = datetime.strptime(data, "%Y-%m-%d").strftime("%d/%m/%Y")
        except ValueError:
            data_formatada = data
        
        response = requests.post(url, json=payload, headers=headers)
        if response.status_code == 200:
            exibir_log(f"Justificativa registrada para matricula {matricula} - Dia: {data_formatada} - Tipo: {descricao} - Qtd. Horas {qtd_horas}")
        else:
            exibir_log(f"Erro ao enviar justificativa para ID {matricula}: {response.status_code}")

def envio_planilha_desligamento():
    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione o arquivo de cadastro de pessoas",
        filetypes=[("Arquivo Excel", "*.xlsx *.xls")]
    )
    if not caminho_arquivo:
        return
    
    try:
        df = pd.read_excel(caminho_arquivo)
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao ler o arquivo: {e}")
        return
    
    
    resultados = []
    df = df.fillna('')
    if api_var.get() == "Kairos":
        url = 'https://www.dimepkairos.com.br/RestServiceApi/Dismiss/MarkDismiss'
    else:
        url = 'https://www.mdcomune.com.br/RestServiceApi/Dismiss/MarkDismiss'
    headers = {
        'identifier': dados_selecionados['CNPJ'],
        'key': dados_selecionados['Chave API'],
        'User-Agent': 'PostmanRuntime/7.30.0'
    }
    
    for index, row in df.iterrows():
        try:
            Matricula = row['Matricula']
            Data_desligamento = row['Data de Desligamento']
            
            
            try:
                if isinstance(Data_desligamento, datetime):
                    data_formatada = Data_desligamento.strftime("%Y-%m-%d")
                else:
                    data_formatada = datetime.strptime(str(Data_desligamento), "%d/%m/%Y").strftime("%Y-%m-%d")
            except ValueError:
                exibir_log(f"Data Inicio invalida para matrícula {Matricula}. Ignorando.")
                continue

            
            payload = {
                "MATRICULA": Matricula,
                "MOTIVO": "11-Rescisão sem justa causa por iniciativa do empregador",
                "DATA": data_formatada
            }
            
            print(payload)
            
            response = requests.post(url, json=payload, headers=headers)
            response_json = response.json()

            if response_json.get("Sucesso"):
                status = "Sucesso"
                mensagem = response_json.get("Mensagem", "")
                exibir_log(f'Desligamento marcado com sucesso para a {Matricula}')
            else:
                status = "Falha"
                mensagem = response_json.get("Mensagem", "Erro desconhecido")

        except Exception as e:
            status = "Erro"
            mensagem = str(e)

        resultados.append([Matricula, status, mensagem])
    if resultados:
        sucesso = len([res for res in resultados if res[1] == "Sucesso"])
        falha = len([res for res in resultados if res[1] != "Sucesso"])
        exibir_log(f"Processo finalizado. Sucessos: {sucesso}, Falhas: {falha}")
        messagebox.showinfo(
        "Resumo do Cadastro",
        f"Processo concluído:\nSucessos: {sucesso}\nFalhas: {falha}"
        )

def envio_planilha_ferias():
    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione o arquivo de cadastro de pessoas",
        filetypes=[("Arquivo Excel", "*.xlsx *.xls")]
    )
    if not caminho_arquivo:
        return
    
    try:
        df = pd.read_excel(caminho_arquivo)
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao ler o arquivo: {e}")
        return
    
    
    resultados = []
    df = df.fillna('')
    if api_var.get() == "Kairos":
        url = 'https://www.dimepkairos.com.br/RestServiceApi/Holiday/MarkHoliday'
    else:
        url = 'https://www.mdcomune.com.br/RestServiceApi/Holiday/MarkHoliday'
        
    headers = {
        'identifier': dados_selecionados['CNPJ'],
        'key': dados_selecionados['Chave API'],
        'User-Agent': 'PostmanRuntime/7.30.0'
    }
    
    for index, row in df.iterrows():
        try:
            Matricula = row['Matricula']
            Data_inicio = row['Data Inicio']
            Data_fim = row['Data Fim']
            
            
            try:
                if isinstance(Data_inicio, datetime):
                    data_formatada_inicio = Data_inicio.strftime("%Y-%m-%d")
                else:
                    data_formatada_inicio = datetime.strptime(str(Data_inicio), "%d/%m/%Y").strftime("%Y-%m-%d")
            except ValueError:
                exibir_log(f"Data Inicio invalida para matrícula {Matricula}. Ignorando.")
                continue

            try:
                if isinstance(Data_inicio, datetime):
                    data_formatada_fim = Data_fim.strftime("%Y-%m-%d")
                else:
                    data_formatada_fim = datetime.strptime(str(Data_fim), "%d/%m/%Y").strftime("%Y-%m-%d")
            except ValueError:
                exibir_log(f"Data Final invalida para matrícula {Matricula}. Ignorando.")
                continue
            
            payload = {
                "MATRICULA": Matricula,
                "DATAINICIO": data_formatada_inicio,
                "DATAFIM": data_formatada_fim,
                "ConfirmacaoFeriasPrimeiroPeriodoAquisitivo": 'true',
            }
            
            response = requests.post(url, json=payload, headers=headers)
            response_json = response.json()

            if response_json.get("Sucesso"):
                status = "Sucesso"
                mensagem = response_json.get("Mensagem", "")
                exibir_log(f'Ferias enviadas com SUCESSO para a {Matricula}')
            else:
                status = "Falha"
                mensagem = response_json.get("Mensagem", "Erro desconhecido")

        except Exception as e:
            status = "Erro"
            mensagem = str(e)

        resultados.append([Matricula, status, mensagem])
        
    if resultados:
        sucesso = len([res for res in resultados if res[1] == "Sucesso"])
        falha = len([res for res in resultados if res[1] != "Sucesso"])
        exibir_log(f"Processo finalizado. Sucessos: {sucesso}, Falhas: {falha}")
        messagebox.showinfo(
        "Resumo do Cadastro",
        f"Processo concluído:\nSucessos: {sucesso}\nFalhas: {falha}"
        )

def cadastrar_pessoas():
    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione o arquivo de cadastro de pessoas",
        filetypes=[("Arquivo Excel", "*.xlsx *.xls")]
    )
    if not caminho_arquivo:
        return

    try:
        df = pd.read_excel(caminho_arquivo, dtype={"CPF": str})
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao ler o arquivo: {e}")
        return

    resultados = []
    if api_var.get() == "Kairos":
        url = "https://www.dimepkairos.com.br/RestServiceApi/People/SavePerson"
    else:
        url = "https://www.mdcomune.com.br/RestServiceApi/People/SavePerson"
         
    headers = {
        "identifier": dados_selecionados["CNPJ"],
        "key": dados_selecionados["Chave API"],
        "cpf": dados_selecionados['CPF Responsável'],
        "User-Agent": "PostmanRuntime/7.30.0"
    }
    for _, row in df.iterrows():
        try:
            Matricula = row['Matricula']
            Cracha = row['Cracha']
            Nome = row['Nome Completo']
            Cpf = str(row['CPF']).zfill(11)
            Rg = row.get('RG', None)
            Email = row.get('Email', None)
            DataAdm = pd.to_datetime(row["Admissão"]).strftime('%d-%m-%Y')
            try:
                DataNasc = pd.to_datetime(row["Nascimento"]).strftime('%d-%m-%Y') if pd.notna(row["Nascimento"]) else None
            except Exception:
                DataNasc = None  # Caso a data seja inválida ou ausente
            BaseHoras = row['Base de Horas']
            EstruturaOrg = row['Estrutura']
            Horario = row['Horário']
            RegraCalculo = row['Cálculo']
            Pis = row.get('PIS', None)
            PisAuto = row.get('Possui PIS?', 1)
            Sexo = row['Sexo']
            Cargo = row.get('Cargo', None)
            CampoAlternativo = row.get('Campo Alternativo', None)
            Endereço = row.get('Endereço', None)
            Numero = row.get('Número', None)
            Bairro = row.get('Bairro', None)
            Cidade = row.get('Cidade', None)
            Estado = row.get('Estado', None)
            Pais = row.get('Pais', None)
                        
            if pd.notna(Cargo):
                Cargo = str(int(float(Cargo))) if isinstance(Cargo, (float, int)) else str(Cargo)

            payload = {
                "Matricula": Matricula,
                "Cracha": Cracha,
                "Nome": Nome,
                "Cpf": Cpf,
                "CpfResponsavel": dados_selecionados["CPF Responsável"],
                "DataAdmissao": DataAdm,
                "BaseHoras": BaseHoras,
                "Estrutura": {"Id": EstruturaOrg},
                "TipoFuncionario": {"IdTipoFuncionario": 1},
                "TipoSalario": {"Id": 101},
                "Horarios": [
                    {
                        "Id": 0,
                        "Horario": {"Id": Horario},
                        "Inicio": DataAdm,
                        "Fim": "31/12/9999 00:00:00"
                    }
                ],
                "RegrasCalculo": [
                    {
                        "Id": 0,
                        "Regra": {"Id": RegraCalculo},
                        "Inicio": DataAdm,
                        "Fim": "31/12/9999 00:00:00"
                    }
                ],
                "FlagGerarNumeroPISAutomatico": bool(PisAuto == 0),
                "Sexo": Sexo
            }
            if pd.notna(Endereço):
                payload["Rua"] = Endereço
                
            if pd.notna(Numero):
                payload["NumeroAndar"] = Numero
                
            if pd.notna(Bairro):
                payload["Bairro"] = Bairro
            
            if pd.notna(Cidade):
                payload["Cidade"] = Cidade
                
            if pd.notna(Estado):
                payload["Estado"] = Estado
                
            if pd.notna(Pais):
                payload["Pais"] = Pais
            
            if pd.notna(Rg):
                payload["Rg"] = DataNasc
            
            if pd.notna(DataNasc):
                payload["DataNascimento"] = DataNasc

            if pd.notna(Email):
                payload["Email"] = str(Email).strip()

            if pd.notna(CampoAlternativo):
                payload["CampoAlternativo1"] = str(CampoAlternativo).strip()

            if pd.notna(Cargo):
                payload['Cargo'] = {"Id": Cargo}

            if PisAuto == 1 and Pis:
                payload["CodigoPis"] = Pis

            response = requests.post(url, json=payload, headers=headers)

            try:
                response_json = response.json()
            except json.JSONDecodeError:
                raise ValueError(f"Resposta inválida da API: {response.text}")
            
            if response_json.get("Sucesso"):
                status = "Sucesso"
                mensagem = response_json.get("Mensagem", "")
            else:
                status = "Falha"
                mensagem = response_json.get("Mensagem", "Erro desconhecido")

        except Exception as e:
            status = "Erro"
            mensagem = str(e)
            
        resultados.append([Matricula, status, mensagem])
        exibir_log(f"Matricula: {Matricula} | Status: {status} | Mensagem: {mensagem}")
    if resultados:
        sucesso = len([res for res in resultados if res[1] == "Sucesso"])
        falha = len([res for res in resultados if res[1] != "Sucesso"])
        exibir_log(f"Processo finalizado. Sucessos: {sucesso}, Falhas: {falha}")
        messagebox.showinfo(
        "Resumo do Cadastro",
        f"Processo concluído:\nSucessos: {sucesso}\nFalhas: {falha}"
        )

def alteracao_pessoas_envio():
    
    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione o arquivo de cadastro de pessoas",
        filetypes=[("Arquivo Excel", "*.xlsx *.xls")]
    )
    if not caminho_arquivo:
        return

    try:
        df = pd.read_excel(caminho_arquivo, dtype={"People_Cpf": str})
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao ler o arquivo: {e}")
        return

    resultados = []
    df = df.fillna('')
    if api_var.get() == "Kairos":
        url = 'https://www.dimepkairos.com.br/RestServiceApi/People/ChangePerson'
    else:
        url = 'https://www.mdcomune.com.br/RestServiceApi/People/ChangePerson'
    headers = {
        'identifier': dados_selecionados['CNPJ'],
        'key': dados_selecionados['Chave API'],
        'User-Agent': 'PostmanRuntime/7.30.0'
    }

    for index, row in df.iterrows():
        try:
            Id = row['People_Id']
            Matricula = row['People_Matricula']
            Cracha = row['People_Cracha']
            Nome = row['People_Nome']
            DataAdm = row['People_DataAdmissao']
            Rg = row.get('People_Rg', ' ')
            Cpf = str(row['People_Cpf'].replace(".","").replace("-",""))

            if 'People_Email' in df.columns:
                Email = row['People_Email']
                Email = str(Email).strip() if pd.notna(Email) else None
                if Email == '':
                    Email = None
            else:
                Email = None

            BaseHoras = row['People_BaseHoras']
            Pis = row['People_CodigoPis']
            Sexo = row['People_Sexo']
            EstruturaId = row['People_Estrutura.Id']
            TipoSalario = row['People_TipoSalario.Id']
            Cargo = row['People_Cargo.Id']

            Pis = None if pd.isna(Pis) or Pis == '' else str(Pis)

            if Pis is None:
                payload = {
                    "Id": Id,
                    "Matricula": Matricula,
                    "Cracha": Cracha,
                    "Nome": Nome,
                    "DataAdmissao": DataAdm,
                    "Rg": Rg,
                    "Cpf": Cpf,
                    "BaseHoras": BaseHoras,
                    "Estrutura": {"Id": EstruturaId},
                    "TipoFuncionario": {"Id": 1},
                    "TipoSalario": {"Id": 101},
                    "Sexo": Sexo,
                    "FlagGerarNumeroPISAutomatico": True,
                    "CpfResponsavel": dados_selecionados["CPF Responsável"]
                }
            else:
                payload = {
                    "Id": Id,
                    "Matricula": Matricula,
                    "Cracha": Cracha,
                    "Nome": Nome,
                    "DataAdmissao": DataAdm,
                    "Rg": Rg,
                    "Cpf": Cpf,
                    "BaseHoras": BaseHoras,
                    "Estrutura": {"Id": EstruturaId},
                    "TipoFuncionario": {"Id": 1},
                    "TipoSalario": {"Id": 101},
                    "Sexo": Sexo,
                    "FlagGerarNumeroPISAutomatico": False,
                    "Pis": Pis,
                    "CpfResponsavel": dados_selecionados["CPF Responsável"]
                }

            if Email is not None:
                payload["Email"] = Email

            print(payload)
            response = requests.post(url, json=payload, headers=headers)
            response_json = response.json()

            if response_json.get("Sucesso"):
                status = "Sucesso"
                mensagem = response_json.get("Mensagem", "")
            else:
                status = "Falha"
                mensagem = response_json.get("Mensagem", "Erro desconhecido")

        except Exception as e:
            status = "Erro"
            mensagem = str(e)

        resultados.append([Matricula, status, mensagem])
        
        exibir_log(f"Matricula: {Matricula} | Status: {status} | Mensagem: {mensagem}")
    if resultados:
        sucesso = len([res for res in resultados if res[1] == "Sucesso"])
        falha = len([res for res in resultados if res[1] != "Sucesso"])
        exibir_log(f"Processo finalizado. Sucessos: {sucesso}, Falhas: {falha}")
        messagebox.showinfo(
        "Resumo do Cadastro",
        f"Processo concluído:\nSucessos: {sucesso}\nFalhas: {falha}"
        )

def cadastro_cargo():
    """Cadastra cargos na API usando dados de um arquivo Excel."""
    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione o arquivo de cadastro de cargos",
        filetypes=[("Arquivo Excel", "*.xlsx *.xls")]
    )
    if not caminho_arquivo:
        exibir_log("Nenhum arquivo selecionado.")
        return

    try:
        df = pd.read_excel(caminho_arquivo)
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao ler o arquivo: {e}")
        exibir_log(f"Erro ao carregar o arquivo: {e}")
        return

    # Verifica se as colunas necessárias estão presentes
    colunas_necessarias = ["Codigo(Novo Cargo)", "Descricao(Novo Cargo)"]
    if not all(col in df.columns for col in colunas_necessarias):
        messagebox.showerror(
            "Erro", f"O arquivo deve conter as colunas: {', '.join(colunas_necessarias)}"
        )
        exibir_log(f"Arquivo inválido. Colunas esperadas: {colunas_necessarias}")
        return

    resultados = []
    df = df.fillna("")  # Substitui valores NaN por strings vazias
    if api_var.get() == "Kairos":
        url = "https://www.dimepkairos.com.br/RestServiceApi/JobPosition/SaveJobPosition"
    else:
        url = "https://www.mdcomune.com.br/RestServiceApi/JobPosition/SaveJobPosition"
    headers = {
        "identifier": dados_selecionados["CNPJ"],
        "key": dados_selecionados["Chave API"],
        "User-Agent": "PostmanRuntime/7.30.0"
    }

    for index, row in df.iterrows():
        try:
            codigo = row["Codigo(Novo Cargo)"]
            descricao = row["Descricao(Novo Cargo)"]
            
            if pd.notna(codigo):
                codigo = str(int(float(codigo))) if isinstance(codigo, (float, int)) else str(codigo)

            # Ignorar linhas com valores obrigatórios faltando
            if not codigo or not descricao:
                exibir_log(f"Linha {index + 1} ignorada: Código ou Descrição ausente.")
                continue

            payload = {
                "Codigo": codigo,
                "Descricao": str(descricao).strip()
            }
            print (payload)
            response = requests.post(url, json=payload, headers=headers)

            # Verifica a resposta da API
            if response.status_code == 200:
                response_json = response.json()
                if response_json.get("Sucesso"):
                    status = "Sucesso"
                    mensagem = response_json.get("Mensagem", "Cadastro realizado com sucesso.")
                    exibir_log(f"Cargo '{descricao}' cadastrado com sucesso.")
                else:
                    status = "Falha"
                    mensagem = response_json.get("Mensagem", "Erro desconhecido.")
                    exibir_log(f"Falha ao cadastrar cargo '{descricao}': {mensagem}")
            else:
                status = "Erro"
                mensagem = f"Erro de comunicação com a API: {response.status_code}"
                exibir_log(f"Erro ao cadastrar cargo '{descricao}': {mensagem}")

        except Exception as e:
            status = "Erro"
            mensagem = str(e)
            exibir_log(f"Erro ao processar linha {index + 1}: {mensagem}")

        # Armazena o resultado da linha
        resultados.append([descricao, status, mensagem])

    # Mostra um resumo dos resultados
    if resultados:
        sucesso = len([res for res in resultados if res[1] == "Sucesso"])
        falha = len([res for res in resultados if res[1] != "Sucesso"])
        exibir_log(f"Processo finalizado. Sucessos: {sucesso}, Falhas: {falha}")
        messagebox.showinfo(
            "Resumo do Cadastro",
            f"Processo concluído:\nSucessos: {sucesso}\nFalhas: {falha}"
        )

#Funções Auxiliares

def salvar_dataframe(df, nome):
    output_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel file", "*.xlsx")],
        initialfile=f'{nome}.xlsx'
    )
    if output_path:
        df.to_excel(output_path, index=False)
        exibir_log(f"Arquivo {nome} salvo com Sucesso")

def aplicar_estilo(output_path):
    wb = load_workbook(output_path)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    fill_grey = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.row % 2 == 0:
                cell.fill = fill_grey
            else:
                cell.fill = fill_white
            cell.alignment = Alignment(horizontal="left", vertical="center")

    apply_borders(ws)

    wb.save(output_path)

def iniciar_coleta(selecoes, popup):
    """Inicia a coleta com base nas seleções."""
    popup.destroy()

    for nome, (var, func) in selecoes.items():
        if var.get():
            exibir_log(f"Função: {nome} sendo realizada para Empresa: {dados_selecionados['Razão Social']}")
            func()

    messagebox.showinfo("Concluido", "Processo de coleta realizado com sucesso")
    
def iniciar_envio(selecoes, popup):
    """Inicia a coleta com base nas seleções."""
    popup.destroy()

    for nome, (var, func) in selecoes.items():
        if var.get():
            exibir_log(f"Função: {nome} sendo realizada para Empresa: {dados_selecionados['Razão Social']}")
            func()

def apply_borders(ws):
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

def generate_date_range(start_date, end_date):
    return pd.date_range(start=start_date, end=end_date)

def clean_json_date(date_str):
    return date_str[:10]

def process_faltas(df):
    faltas = pd.DataFrame(columns=df.columns)

    for index, row in df.iterrows():
        apontamentos = row['Apontamentos']
        descontos = row['Descontos']
        horario = row['Horario']
        
        if 'Falta' in str(apontamentos) or 'Atraso' in str(apontamentos) or \
           'Falta' in str(descontos) or 'Atraso' in str(descontos):
            faltas = pd.concat([faltas, row.to_frame().T])
        
        elif pd.isna(apontamentos) and pd.notna(horario):
            faltas = pd.concat([faltas, row.to_frame().T])

    faltas['Data'] = pd.to_datetime(faltas['Data'], format='%d/%m/%Y').dt.strftime('%d/%m/%Y')

    return faltas

def processar_marcacoes_com_cpf():
    """Processa um arquivo JSON contendo marcações de CPF, busca as credenciais e envia as marcações."""
    # Seleção do arquivo
    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione o arquivo de marcações (JSON)",
        filetypes=[("Arquivos JSON", "*.json")]
    )
    if not caminho_arquivo:
        print("Nenhum arquivo selecionado.")
        return

    # Leitura do arquivo JSON
    try:
        with open(caminho_arquivo, "r", encoding="utf-8") as file:
            dados = json.load(file)
    except Exception as e:
        print(f"Erro ao abrir o arquivo JSON: {e}")
        return

    # Verificar estrutura do JSON
    registros = dados.get("records", None)
    if not registros or not isinstance(registros, list):
        print("Formato do JSON inválido. Esperado uma lista dentro da chave 'records'.")
        return

    # Configuração da API
    if api_var.get() == "Kairos":
        url_search = "https://www.dimepkairos.com.br/RestServiceApi/People/SearchPeople"
        url_punch = "https://www.dimepkairos.com.br/RestServiceApi/Mark/SetMarks"
    else:
        url_search = "https://www.mdcomune.com.br/RestServiceApi/People/SearchPeople"
        url_punch = "https://www.mdcomune.com.br/RestServiceApi/Mark/SetMarks"
        
    headers = {
        "identifier": dados_selecionados["CNPJ"],
        "key": dados_selecionados["Chave API"],
        "User-Agent": "PostmanRuntime/7.30.0"
    }

    # Processar cada marcação
    for registro in registros:
        cpf = registro.get("Cpf", "").zfill(11).replace(".", "").replace("-", "")
        dt_hr = registro.get("DtHr", "")

        # Validação básica
        if not cpf or not dt_hr:
            print(f"Dados incompletos: CPF={cpf}, DtHr={dt_hr}")
            continue

        # Formatar DataHoraApontamento
        try:
            dt_hr_formatado = datetime.strptime(dt_hr, "%d/%m/%Y %H:%M:%S").strftime("%d/%m/%Y %H:%M")
        except ValueError:
            print(f"DataHora inválida: {dt_hr} para CPF {cpf}")
            continue

        # Buscar matrícula pelo CPF
        try:
            payload_search = {"Cpf": cpf}
            response = requests.post(url_search, json=payload_search, headers=headers)
            response_data = response.json()

            if not response_data.get("Sucesso", False):
                print(f"Erro ao buscar CPF {cpf}: {response_data.get('Mensagem', 'Erro desconhecido')}")
                continue

            # Encontrar matrícula associada ao CPF
            pessoas = response_data.get("Obj", [])
            matricula_encontrada = None
            for pessoa in pessoas:
                if pessoa.get("Cpf", "").replace(".", "").replace("-", "") == cpf:
                    matricula_encontrada = pessoa.get("Matricula", "")
                    break

            if not matricula_encontrada:
                print(f"Matrícula não encontrada para CPF {cpf}")
                continue

            # Enviar marcação usando a matrícula e DataHora do arquivo JSON
            payload_punch = {
                "Matricula": matricula_encontrada,
                "DataHoraApontamento": dt_hr_formatado,
                "CpfResponsavel": dados_selecionados["CPF Responsável"],
                "ResponseType": "AS400V1"
            }
            punch_response = requests.post(url_punch, json=payload_punch, headers=headers)
            punch_data = punch_response.json()

            if punch_data.get("Sucesso", False):
                print(f"Marcação enviada com sucesso para CPF {cpf}, Matrícula {matricula_encontrada}, DataHora: {dt_hr_formatado}")
            else:
                print(f"Erro ao enviar marcação para CPF {cpf}: {punch_data.get('Mensagem', 'Erro desconhecido')}")

        except Exception as e:
            print(f"Erro ao processar CPF {cpf}: {e}")
    
def filtra_marcacoes_impares_e_htrab_vazio(entradas):
    marcacoes_filtradas = []

    for entrada in entradas:
        apontamentos = entrada.get("Apontamentos", "").strip()
        htrab = entrada.get("HTrab", "").strip()
        
        if not htrab and apontamentos:
            intervals = apontamentos.split()
            has_incomplete_turn = False
            for interval in intervals:
                if '-' not in interval or interval.count('-') == 1 and interval.endswith('-'):
                    has_incomplete_turn = True
                    break
            
            if has_incomplete_turn:
                marcacoes_filtradas.append(entrada)

    return marcacoes_filtradas

def process_incomum(df):
    incomuns = pd.DataFrame(columns=df.columns)

    for index, row in df.iterrows():
        apontamentos = row['Apontamentos']
        
        if not apontamentos:
            continue

        apontamentos_parts = re.findall(r'\d{2}:\d{2}', str(apontamentos))
        horario = row['Horario']
        horario_parts = re.findall(r'\d{2}:\d{2}', str(horario))
        
        if len(apontamentos_parts) > 0 and len(horario_parts) > 0:
            apontamentos_length = len(apontamentos_parts)
            horario_length = len(horario_parts)
            if apontamentos_length % 2 == 0 and apontamentos_length > 0 and apontamentos_length != horario_length:
                incomuns = pd.concat([incomuns, row.to_frame().T])

    incomuns['Data'] = pd.to_datetime(incomuns['Data'], format='%d/%m/%Y').dt.strftime('%d/%m/%Y')
    return incomuns

def carregar_justificativas():
    try:
        with open("justificativas_selecionadas.txt", "r") as f:
            justificativas = f.readlines()
            justificativas = [line.strip() for line in justificativas]
    except FileNotFoundError:
        justificativas = []
    
    return justificativas

def get_data_from_api(url, payload, headers):
    try:
        response = requests.post(url, json=payload, headers=headers)
        response.raise_for_status()

        data = response.json()
        if 'Obj' in data and isinstance(data['Obj'], list):
            obj_data = pd.json_normalize(data['Obj'])
            return obj_data

        elif 'Obj' in data and isinstance(data['Obj'], str):
            obj_data = json.loads(data['Obj'])
            return pd.json_normalize(obj_data)
        else:
            return pd.json_normalize(data)

    except ValueError as e:
        print(f"Erro ao decodificar JSON: {e}")
        print("Conteúdo da resposta:", response.text)
    except requests.exceptions.RequestException as e:
        print(f"Erro na requisição: {e}")

def combinar_data_hora(data_completa, hora, tipo_marcacao="", horario_intervalo=None):
    if pd.notna(data_completa) and pd.notna(hora):
        if isinstance(data_completa, pd.Timestamp):
            data = data_completa.strftime("%d/%m/%Y")
        else:
            data = str(data_completa)

        hora_str = hora if isinstance(hora, str) else hora.strftime("%H:%M")
        if len(hora_str) > 5:
            hora_str = hora_str[:5]

        data_hora_str = f"{data} {hora_str}"
        try:
            data_hora_dt = datetime.strptime(data_hora_str, "%d/%m/%Y %H:%M")

            # Determinar se a marcação é noturna com base no(s) intervalo(s)
            if horario_intervalo:
                try:
                    # Dividir o intervalo em partes (caso contenha "|")
                    intervalos = [intervalo.strip() for intervalo in horario_intervalo.split("|")]

                    # Extrair o primeiro horário do primeiro intervalo para validação base
                    primeiro_horario = datetime.strptime(intervalos[0].split(" - ")[0], "%H:%M").time()

                    for intervalo in intervalos:
                        inicio, fim = intervalo.split(" - ")
                        inicio_hora = datetime.strptime(inicio, "%H:%M").time()
                        fim_hora = datetime.strptime(fim, "%H:%M").time()

                        # Ajustar para dia seguinte se a marcação estiver no intervalo ou passou do limite
                        if data_hora_dt.time() >= inicio_hora or data_hora_dt.time() <= fim_hora:
                            # Ajuste com base no primeiro horário
                            if data_hora_dt.time() < primeiro_horario:
                                data_hora_dt += timedelta(days=1)
                            break  # Já encontrou o intervalo correto, pode parar

                except ValueError as e:
                    print(f"Erro ao interpretar intervalo de horário: {e}")

            return data_hora_dt.strftime("%d/%m/%Y %H:%M")

        except ValueError:
            try:
                data_hora_dt = datetime.strptime(data_hora_str, "%d/%m/%Y %H:%M:%S")
                return data_hora_dt.strftime("%d/%m/%Y %H:%M")
            except ValueError as e:
                print(f"Erro ao converter data e hora: {e}")
                return None
    return None

if not os.path.exists("justificativas"):
    os.makedirs("justificativas")
    
def funcao_justificativa_get():

    if not dados_selecionados.get('CNPJ') or not dados_selecionados.get('Chave API'):
        messagebox.showerror("Erro", "Nenhuma empresa selecionada. Selecione uma empresa antes de carregar justificativas.")
        return

    # Substituir caracteres inválidos no nome do arquivo
    cnpj = dados_selecionados['CNPJ'].replace("/", "-")
    arquivo_justificativas = os.path.join("justificativas", f"{cnpj}_justificativas.txt")

    def carregar_justificativas():
        try:
            payload = {"Code": 0, "IdType": 1202, "ResponseType": "AS400V1"}
            headers = {
                "identifier": dados_selecionados['CNPJ'],
                "key": dados_selecionados['Chave API'],
                'User-Agent': 'PostmanRuntime/7.30.0'
            }
            if api_var.get() == "Kairos":
                url = 'https://www.dimepkairos.com.br/RestServiceApi/Justification/GetJustification'
            else:
                url = 'https://www.mdcomune.com.br/RestServiceApi/Justification/GetJustification'
                
            response = requests.post(url, json=payload, headers=headers)
            response.raise_for_status()
            data = response.json()

            justificativas = [
                f"{item['Id']}|{item['Description']}"
                for item in data.get("Obj", [])
            ]

            for widget in frame_checkboxes.winfo_children():
                widget.destroy()

            for linha in justificativas:
                try:
                    id_justificativa, descricao = linha.strip().split("|", 1)
                    var = tk.BooleanVar()
                    checkbox = ttk.Checkbutton(
                        frame_checkboxes,
                        text=f"{descricao} (ID: {id_justificativa})",
                        variable=var
                    )
                    checkbox.grid(sticky="w", padx=10, pady=5)
                    justificativas_selecionadas[int(id_justificativa)] = (var, descricao)
                except ValueError:
                    continue

        except requests.RequestException as e:
            messagebox.showerror("Erro", f"Erro ao carregar justificativas: {e}")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro inesperado: {e}")

    def gravar_justificativas():
        selecionadas = [
            f"{id_justificativa}|{descricao}"
            for id_justificativa, (var, descricao) in justificativas_selecionadas.items() if var.get()
        ]

        if not selecionadas:
            messagebox.showwarning("Atenção", "Nenhuma justificativa selecionada.")
            return

        if not os.path.exists("justificativas"):
            os.makedirs("justificativas")

        with open(arquivo_justificativas, "w", encoding="utf-8") as f:
            f.write("\n".join(selecionadas))

        messagebox.showinfo("Sucesso", "Justificativas selecionadas foram gravadas.")
        popup_futuro.destroy()

    estilo_botao = "success" if api_var.get() == "Kairos" else "info"
    popup_futuro = tk.Toplevel(root)
    popup_futuro.title("Seleção de Justificativas")
    popup_futuro.geometry("400x380")
    popup_futuro.iconbitmap("Mantis.ico")

    canvas = tk.Canvas(popup_futuro)
    scrollbar = ttk.Scrollbar(popup_futuro, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=scrollbar.set)

    frame_checkboxes = ttk.Frame(canvas)
    canvas.create_window((0, 0), window=frame_checkboxes, anchor="nw")

    def on_mouse_wheel(event):
        canvas.yview_scroll(-1 * (event.delta // 120), "units")

    canvas.bind("<MouseWheel>", on_mouse_wheel)

    scrollbar.pack(side="right", fill="y")
    canvas.pack(fill="both", expand=True)

    justificativas_selecionadas = {}

    carregar_justificativas()

    frame_checkboxes.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))

    ttk.Button(
        popup_futuro, text="Gravar Seleção", command=gravar_justificativas, bootstyle=estilo_botao
    ).pack(pady=10)

    ttk.Button(
        popup_futuro, text="Fechar", command=popup_futuro.destroy, bootstyle="danger"
    ).pack(pady=10)

def carregar_justificativas_salvas(cnpj):
    # Substituir barras por hífens no CNPJ para evitar problemas de nomes de arquivo
    cnpj_formatado = cnpj.replace("/", "-")
    arquivo_justificativas = os.path.join("justificativas", f"{cnpj_formatado}_justificativas.txt")

    if not os.path.exists(arquivo_justificativas):
        exibir_log(f"Arquivo de justificativas não encontrado para o CNPJ: {cnpj}")
        return []

    try:
        with open(arquivo_justificativas, "r", encoding="utf-8") as f:
            linhas = f.readlines()
        justificativas = [linha.strip().split("|")[1] for linha in linhas if "|" in linha]
        return justificativas
    except Exception as e:
        exibir_log(f"Erro ao carregar justificativas: {e}")
        return []

def selecionar_arquivo_empresas():
    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione o arquivo Empresas",
        filetypes=[("Arquivos Excel", "*.xlsx")]
    )
    if not caminho_arquivo:
        messagebox.showwarning("Aviso", "Nenhum arquivo selecionado.")
        return
    
    try:
        global df_empresas
        df_empresas = pd.read_excel(caminho_arquivo, dtype={"CPF Responsável": str})

        combo_razao_social['values'] = df_empresas["Razão Social"].tolist()
        combo_razao_social.set('')
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao carregar o arquivo: {e}")
        return

def carregar_ids_justificativas():
    ids_justificativas = {}
    with open("justificativas_ids_descricoes.txt", "r") as file:
        for line in file:
            id_just, descricao = line.strip().split(",")
            ids_justificativas[descricao] = int(id_just)
    return ids_justificativas

def buscar_dados_funcionario(matricula):
    """Busca CPF e ID do funcionário pela matrícula."""
    
    if api_var.get() == "Kairos":
        url = "https://www.dimepkairos.com.br/RestServiceApi/People/SearchPerson"
    else:
        url = "https://www.mdcomune.com.br/RestServiceApi/People/SearchPerson"
        
    headers = {
        'identifier': dados_selecionados["CNPJ"],
        'key': dados_selecionados["Chave API"],
        'User-Agent': 'PostmanRuntime/7.30.0'
    }
    payload = {"Matricula": matricula}

    response = requests.post(url, json=payload, headers=headers)

    if response.status_code == 200:
        try:
            data_response = response.json()
            funcionarios = json.loads(data_response['Obj']) if isinstance(data_response['Obj'], str) else data_response['Obj']
            funcionario = funcionarios[0]
            return funcionario.get('Cpf'), funcionario.get('Id')
        except (json.JSONDecodeError, IndexError, KeyError):
            exibir_log(f"Erro ao processar resposta da API para matrícula {matricula}.")
    else:
        exibir_log(f"Erro na API ao buscar dados do funcionário {matricula}: {response.status_code}")
    return None, None

def validar_horas(qtd_horas):
    """Valida e formata as horas."""
    if not qtd_horas or str(qtd_horas).strip() == "":
        return "12:00"
    try:
        return datetime.strptime(str(qtd_horas), "%H:%M:%S").strftime("%H:%M")
    except ValueError:
        try:
            return datetime.strptime(str(qtd_horas), "%H:%M").strftime("%H:%M")
        except ValueError:
            return "12:00"

def carregar_justificativas_salvas_envio(cnpj):
    cnpj_formatado = cnpj.replace("/", "-")
    arquivo_justificativas = os.path.join("justificativas", f"{cnpj_formatado}_justificativas.txt")
    if not os.path.exists(arquivo_justificativas):
        exibir_log(f"Arquivo de justificativas não encontrado para o CNPJ {cnpj}.")
        return {}

    ids_justificativas = {}
    try:
        with open(arquivo_justificativas, "r", encoding="utf-8") as f:
            for linha in f:
                try:
                    id_justificativa, descricao = linha.strip().split("|", 1)
                    ids_justificativas[descricao] = int(id_justificativa)
                except ValueError:
                    continue
    except Exception as e:
        exibir_log(f"Erro ao carregar justificativas salvas: {e}")
        return {}

    return ids_justificativas

def processar_arquivo_excel(df, barra=None, label_status=None, cancelar=None):
    if df.empty:
        exibir_log("Arquivo vazio. Operação abortada.")
        return

    # Carrega o arquivo de justificativas do CNPJ
    cnpj = dados_selecionados['CNPJ']
    ids_justificativas = carregar_justificativas_salvas_envio(cnpj)

    if not ids_justificativas:
        exibir_log(f"Justificativas não encontradas para o CNPJ {cnpj}. Carregue as justificativas primeiro.")
        return

    # Verifica se há justificativas do tipo "Folga" e prepara o arquivo
    tem_folga = any(row.get('Justificativa') == "Folga" for _, row in df.iterrows())
    caminho_folgas = None

    if tem_folga:
        caminho_folgas = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Arquivo de Texto", "*.txt")],
            title="Selecione o local para salvar o arquivo de folgas"
        )
        if not caminho_folgas:
            exibir_log("Nenhum local de salvamento selecionado para o arquivo de folgas. Ignorando folgas.")
            tem_folga = False

    if tem_folga and caminho_folgas:
        with open(caminho_folgas, "w") as arquivo_folgas:
            for _, row in df.iterrows():
                if cancelar and cancelar.is_set():
                    exibir_log("Envio de justificativas cancelado pelo usuário.")
                    return

                matricula = row.get('Matricula')
                descricao_justificativa = row.get('Justificativa')
                data = row.get('Data')
                pis = row.get('PIS')

                if descricao_justificativa == "Folga":
                    try:
                        data_formatada = datetime.strptime(str(data), "%d/%m/%Y").strftime("%Y-%m-%d")
                    except ValueError:
                        exibir_log(f"Data inválida para matrícula {matricula}. Ignorando.")
                        continue

                    cpf, _ = buscar_dados_funcionario(matricula)
                    if not cpf:
                        exibir_log(f"CPF ausente para matrícula {matricula}. Ignorando linha.")
                        continue

                    linha_folga = (
                        f"{str(matricula).ljust(16)}"
                        f"{str(pis).ljust(23)}"
                        f"{str(data).ljust(12)}"
                        f"{str(cpf).ljust(11)}\n"
                    )
                    arquivo_folgas.write(linha_folga)
                    exibir_log(f"Folga registrada para matrícula {matricula} no dia {data}.")

    # Processa as demais justificativas
    for _, row in df.iterrows():
        if cancelar and cancelar.is_set():
            exibir_log("Envio de justificativas cancelado pelo usuário.")
            return

        matricula = row.get('Matricula')
        descricao_justificativa = row.get('Justificativa')
        data = row.get('Data')
        qtd_horas = row.get('Qtd Horas')

        # Ignorar justificativas inválidas ou folgas (já processadas acima)
        if not descricao_justificativa or pd.isna(descricao_justificativa) or descricao_justificativa == "Folga":
            continue

        try:
            data_formatada = datetime.strptime(str(data), "%d/%m/%Y").strftime("%Y-%m-%d")
        except ValueError:
            exibir_log(f"Data inválida para matrícula {matricula}. Ignorando.")
            continue

        # Relaciona a descrição com o ID da justificativa
        id_justificativa = ids_justificativas.get(descricao_justificativa)
        if id_justificativa is None:
            exibir_log(f"Justificativa '{descricao_justificativa}' não encontrada. Ignorando.")
            continue

        qtd_horas = validar_horas(qtd_horas)

        _, id_funcionario = buscar_dados_funcionario(matricula)
        if not id_funcionario:
            exibir_log(f"ID do funcionário ausente para matrícula {matricula}. Ignorando linha.")
            continue

        # Envia a justificativa
        enviar_justificativa(matricula, id_funcionario, id_justificativa, descricao_justificativa, data_formatada, qtd_horas)

        # Atualiza a barra de progresso
        if barra and not cancelar.is_set():
            barra.step(1)
            barra.update_idletasks()
        if label_status and not cancelar.is_set():
            label_status.config(text=f"Justificativa para Matricula: {matricula}")

    if not cancelar.is_set():
        exibir_log("Processamento de justificativas concluído.")
    else:
        exibir_log("Envio de justificativas interrompido.")

#Funções de Interface

def formatar_data(event, widget):
    
    if event.keysym in ("Left", "Right", "Up", "Down", "BackSpace", "Delete"):
        return  # Não altera o comportamento para essas teclas
    
    texto = widget.entry.get().replace("/", "")  # Remove barras para reformatar
    cursor_pos = widget.entry.index(tk.INSERT)  # Posição atual do cursor
    if len(texto) > 8:  # Limita o tamanho
        texto = texto[:8]
    novo_texto = ""
    cursor_offset = 0  # Para ajustar a posição do cursor com as barras 
    
    # Adiciona as barras automaticamente
    if len(texto) >= 2:
        novo_texto += texto[:2] + "/"
        texto = texto[2:]
        if cursor_pos >= 2:  # Ajusta o cursor para a posição da barra
            cursor_offset += 1
    if len(texto) >= 2:
        novo_texto += texto[:2] + "/"
        texto = texto[2:]
        if cursor_pos >= 4:  # Ajusta o cursor para a posição da barra
            cursor_offset += 1
    novo_texto += texto

    widget.entry.delete(0, tk.END)
    widget.entry.insert(0, novo_texto)
    nova_posicao = cursor_pos + cursor_offset
    widget.entry.icursor(nova_posicao if nova_posicao <= len(novo_texto) else len(novo_texto))

def validar_data(event, widget):
    """Valida e formata a data inserida manualmente."""
    texto = widget.entry.get()
    try:
        # Tenta formatar a data no formato correto
        data_formatada = datetime.strptime(texto, "%d/%m/%Y").strftime("%d/%m/%Y")
        widget.entry.delete(0, tk.END)
        widget.entry.insert(0, data_formatada)
    except ValueError:
        # Se inválida, limpa o campo
        widget.entry.delete(0, tk.END)
        widget.entry.insert(0, "Data inválida")

def confirmar_selecao():
    """Confirma a seleção da empresa e salva as credenciais globalmente."""
    global dados_selecionados
    if not razao_social_var.get():
        messagebox.showwarning("Aviso", "Selecione uma empresa.")
        return
    
    cpf_responsavel = cpf_var.get()[:11]
    cpf_var.set(cpf_responsavel)
    
    dados_selecionados = {
        "Razão Social": razao_social_var.get(),
        "CNPJ": cnpj_var.get(),
        "Chave API": chave_var.get(),
        "CPF Responsável": cpf_responsavel,
        "Data Início" : date_entry_inicio.entry.get(),
        "Data fim": date_entry_fim.entry.get()
    }
    messagebox.showinfo("Sucesso", "Informações selecionadas com sucesso!")

def exibir_log(mensagem):
    """Exibe o log de mensagens na interface."""
    log_widget.configure(state="normal")
    log_widget.insert(tk.END, mensagem + "\n")
    log_widget.see(tk.END)
    log_widget.configure(state="disabled")

def preencher_detalhes():
    try:
        empresa_selecionada = razao_social_var.get()

        if not empresa_selecionada:
            messagebox.showwarning("Aviso", "Nenhuma empresa foi selecionada.")
            return

        empresa_detalhes = df_empresas[df_empresas["Razão Social"] == empresa_selecionada]

        if empresa_detalhes.empty:
            messagebox.showwarning("Aviso", "Empresa selecionada não encontrada.")
            return

        empresa_detalhes = empresa_detalhes.iloc[0]

        cnpj_var.set(empresa_detalhes["CNPJ"])
        chave_var.set(empresa_detalhes["Chave API"])

        cpf_responsavel = str(empresa_detalhes.get("CPF Responsável", "")).zfill(11)
        cpf_var.set(cpf_responsavel)
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao preencher os detalhes: {e}")

def abrir_popup_selecao_pessoas():
    estilo_botao = "success" if api_var.get() == "Kairos" else "info"
    popup = ttk.Toplevel(root)
    popup.title("Selecione a função de envio")
    popup.geometry("500x400")
    popup.iconbitmap("Mantis.ico")
    
    frame_popup = ttk.Frame(popup)
    frame_popup.pack(expand=True, fill="both", padx=10, pady=10)
  
    frame_pessoas = ttk.Labelframe(frame_popup, text="Processo de Cadastro/Alteração", padding=(10, 10))
    frame_pessoas.pack(fill="both", expand=True, padx=10, pady=5)

    # Frame para Coletas de Cadastro e Pessoas
    frame_envio = ttk.Labelframe(frame_popup, text="Alterações Gerais", padding=(10, 10))
    frame_envio.pack(fill="both", expand=True, padx=10, pady=5)

    selecoes_pessoas = {
        "Cadastrar Pessoas": [tk.BooleanVar(), cadastrar_pessoas],
        "Alteração de Pessoas": [tk.BooleanVar(), alteracao_pessoas_envio],
        
    }
    
    selecoes_envio = {
        "Marcar Ferias": [tk.BooleanVar(), envio_planilha_ferias],
        "Marcar Desligamento": [tk.BooleanVar(), envio_planilha_desligamento],
        "Envio Marcação D-REP (JSON)": [tk.BooleanVar(), processar_marcacoes_com_cpf],
        "Cadastrar Cargos": [tk.BooleanVar(), cadastro_cargo]
    }

    for i, (nome, (var, _)) in enumerate(selecoes_pessoas.items()):
        ttk.Checkbutton(frame_pessoas, text=nome, variable=var).grid(row=i, column=0, sticky="w", padx=10, pady=5)

    # Adicionar checkbuttons para as coletas de cadastro e pessoas
    for i, (nome, (var, _)) in enumerate(selecoes_envio.items()):
        ttk.Checkbutton(frame_envio, text=nome, variable=var).grid(row=i, column=0, sticky="w", padx=10, pady=5)

    # Botão de iniciar coleta
    ttk.Button(frame_popup, text="Iniciar Processo", command=lambda: iniciar_envio({**selecoes_pessoas, **selecoes_envio}, popup), bootstyle=estilo_botao
        ).pack(pady=10)
    
    # Configurar o fechamento do popup
    popup.protocol("WM_DELETE_WINDOW", popup.destroy)

def abrir_popup_selecao_coleta():
    estilo_botao = "success" if api_var.get() == "Kairos" else "info"
    popup = ttk.Toplevel(root)
    popup.title("Selecione a planilha a coletar")
    popup.geometry("500x400")
    popup.iconbitmap("Mantis.ico")
    
    # Frame principal
    frame_popup = ttk.Frame(popup)
    frame_popup.pack(expand=True, fill="both", padx=10, pady=10)

    # Frame para Coletas de Planilhas
    frame_planilhas = ttk.Labelframe(frame_popup, text="Coletas de Planilhas", padding=(10, 10))
    frame_planilhas.pack(fill="both", expand=True, padx=10, pady=5)

    # Frame para Coletas de Cadastro e Pessoas
    frame_cadastro = ttk.Labelframe(frame_popup, text="Coletas de Cadastro e Pessoas", padding=(10, 10))
    frame_cadastro.pack(fill="both", expand=True, padx=10, pady=5)

    selecoes_planilhas = {
        "Planilha de Marcação": [tk.BooleanVar(), coleta_planilha_marcacoes],
        "Planilha de Inconsistência": [tk.BooleanVar(), coleta_planilha_marcacoes_inconsistencia],
        "Planilha de Incomum": [tk.BooleanVar(), coleta_planilha_marcacoes_incomum],
        "Planilha de Horas Faltas": [tk.BooleanVar(), coleta_planilha_marcacoes_faltantes],
    }

    selecoes_cadastro = {
        "Coletar Plan. Cadastro": [tk.BooleanVar(), coleta_empresa],
        "Coletar Pessoas Para Alt.": [tk.BooleanVar(), alteracao_pessoas],
        "Coletar Pessoas Para Ferias": [tk.BooleanVar(), coleta_planilha_ferias],
        "Coletar Pessoas para Desligamento": [tk.BooleanVar(), coleta_planilha_desligamento],
        "Coletar Plan. Cargos": [tk.BooleanVar(), coleta_cargos]
    }

    # Adicionar checkbuttons para as coletas de planilhas
    for i, (nome, (var, _)) in enumerate(selecoes_planilhas.items()):
        ttk.Checkbutton(frame_planilhas, text=nome, variable=var).grid(row=i, column=0, sticky="w", padx=10, pady=5)

    # Adicionar checkbuttons para as coletas de cadastro e pessoas
    for i, (nome, (var, _)) in enumerate(selecoes_cadastro.items()):
        ttk.Checkbutton(frame_cadastro, text=nome, variable=var).grid(row=i, column=0, sticky="w", padx=10, pady=5)

    # Botão de iniciar coleta
    ttk.Button(
        frame_popup, text="Iniciar Processo", 
        command=lambda: iniciar_coleta({**selecoes_planilhas, **selecoes_cadastro}, popup), bootstyle=estilo_botao
    ).pack(pady=10)

    # Configurar o fechamento do popup
    popup.protocol("WM_DELETE_WINDOW", popup.destroy)

def mostrar_duas_barras_progresso_paralelo(funcao1, funcao2, df, titulo="Processando"):
    estilo_barra = "success" if api_var.get() == "Kairos" else "info"
    janela = tk.Toplevel()
    janela.title(titulo)
    janela.geometry("400x250")
    janela.resizable(False, False)
    janela.iconbitmap("Mantis.ico")

    cancelar = threading.Event()
    
    df_marcacoes_validas = df[df[['Entrada', 'Almoço Ida', 'Almoço Volta', 'Saida']].notna().any(axis=1)]
    df_justificativas_validas = df[df['Justificativa'].notna()]

    label_status1 = tk.Label(janela, text="Iniciando envio de marcações...", font=("Arial", 12))
    label_status1.pack(pady=5)

    barra1 = ttk.Progressbar(janela, orient="horizontal", bootstyle=estilo_barra, mode="determinate", length=300)
    barra1.pack(pady=5)
    
    label_status2 = tk.Label(janela, text="Iniciando envio de justificativas...", font=("Arial", 12))
    label_status2.pack(pady=5)

    barra2 = ttk.Progressbar(janela, orient="horizontal", bootstyle=estilo_barra, mode="determinate", length=300)
    barra2.pack(pady=5)

    botao_cancelar = ttk.Button(janela, text="Cancelar", bootstyle='danger',command=cancelar.set)
    botao_cancelar.pack(pady=10)

    def executar_tarefa(funcao, barra, label_status, nome_tarefa, df_valido):
        """
        Executa uma função enquanto atualiza a barra e o rótulo correspondentes.
        Respeita o sinalizador de cancelamento.
        """
        try:
            barra["maximum"] = len(df_valido)
            funcao(df_valido, barra, label_status, cancelar)
        except Exception as e:
            exibir_log(f"Erro durante {nome_tarefa}: {e}")
            label_status.config(text=f"Erro: {e}")
        finally:
            if not cancelar.is_set():
                label_status.config(text=f"{nome_tarefa} concluído!")

    thread1 = threading.Thread(target=executar_tarefa, args=(processar_marcacoes, barra1, label_status1, "Envio de Marcações", df_marcacoes_validas))
    thread2 = threading.Thread(target=executar_tarefa, args=(processar_arquivo_excel, barra2, label_status2, "Envio de Justificativas", df_justificativas_validas))

    thread1.start()
    thread2.start()

    def verificar_threads():
        if not thread1.is_alive() and not thread2.is_alive():
            janela.destroy()
        elif cancelar.is_set():
            janela.destroy()
        else:
            janela.after(100, verificar_threads)

    verificar_threads()

    janela.transient()
    janela.grab_set()
    janela.mainloop()

def alterar_tema(sistema):
    """
    Altera a cor dos botões com base no sistema selecionado.
    Kairos -> Verde (success), MD Comune -> Azul (info)
    """
    estilo_botao = "success" if sistema == "Kairos" else "info"

    # Remove os botões antigos
    for widget in botao_frame.winfo_children():
        widget.destroy()

    # Recria os botões com o novo estilo
    botoes_config = [
        ("Tratamento", enviar_dados_combinados),
        ("Pessoas", abrir_popup_selecao_pessoas),
        ("Coleta de Planilhas", abrir_popup_selecao_coleta),
        ("Selecionar Justificativas", funcao_justificativa_get)
    ]
    for i, (texto, comando) in enumerate(botoes_config):
        ttk.Button(botao_frame, text=texto, command=comando, width=botao_width, bootstyle=estilo_botao).grid(
            row=0, column=i, padx=10, pady=5
        )

    global date_entry_inicio, date_entry_fim

    # Remove os widgets antigos
    if "date_entry_inicio" in globals():
        date_entry_inicio.destroy()
    if "date_entry_fim" in globals():
        date_entry_fim.destroy()

    # Recria os widgets DateEntry com o novo estilo
    ttk.Label(frame_campos, text="Data Início:").grid(row=5, column=0, padx=5, pady=5, sticky="w")
    date_entry_inicio = DateEntry(
        frame_campos,
        width=12,
        bootstyle=estilo_botao,
        dateformat="%d/%m/%Y"
    )
    date_entry_inicio.grid(row=5, column=1, padx=5, pady=5, sticky="w")
    date_entry_inicio.entry.bind("<KeyRelease>", lambda e: formatar_data(e, date_entry_inicio))
    date_entry_inicio.entry.bind("<FocusOut>", lambda e: validar_data(e, date_entry_inicio))

    ttk.Label(frame_campos, text="Data Fim:").grid(row=6, column=0, padx=5, pady=5, sticky="w")
    date_entry_fim = DateEntry(
        frame_campos,
        width=12,
        bootstyle=estilo_botao,
        dateformat="%d/%m/%Y"
    )
    date_entry_fim.grid(row=6, column=1, padx=5, pady=5, sticky="w")
    date_entry_fim.entry.bind("<KeyRelease>", lambda e: formatar_data(e, date_entry_fim))
    date_entry_fim.entry.bind("<FocusOut>", lambda e: validar_data(e, date_entry_fim))

    # Log para a mudança de tema
    exibir_log(f"Sistema {sistema} iniciado")

#Interface

root = ttk.Window(themename="darkly")
root.title("Mantis - Kairos")
root.iconbitmap("Mantis.ico")

style = ttk.Style("temamantis-kairos") #Tema Personalidado (Alterar em caso de troca de sistema)

razao_social_var = tk.StringVar()
cnpj_var = tk.StringVar()
chave_var = tk.StringVar()
cpf_var = tk.StringVar()
data_inicio_var = tk.StringVar()
data_fim_var = tk.StringVar()
api_var = tk.StringVar(value="Kairos")  # Variável para alternar entre APIs

frame_selecao = ttk.Frame(root, padding=10)
frame_selecao.grid(row=0, column=0, pady=10, padx=10, sticky="nsew")
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)

frame_campos = ttk.Frame(frame_selecao)
frame_campos.grid(row=0, column=0, pady=10, padx=10)

ttk.Label(root, text="Versão 0.8",).grid(row=1, column=0, padx=5, pady=5, sticky="w")

# Seletor da API
frame_api = ttk.Frame(frame_campos)
frame_api.grid(row=0, column=0, columnspan=4, pady=10)
ttk.Label(frame_campos, text="Selecione o Sistema:").grid(row=0, column=0, padx=5, pady=5, sticky="w")

kairos_button = ttk.Radiobutton(frame_api, text="Kairos", variable=api_var, value="Kairos",
                                command=lambda: alterar_tema("Kairos"))
kairos_button.grid(row=0, column=1, padx=5, pady=5, sticky="w")

mdcomune_button = ttk.Radiobutton(frame_api, text="MD Comune", variable=api_var, value="MDComune",
                                  command=lambda: alterar_tema("MD Comune"))
mdcomune_button.grid(row=0, column=2, padx=5, pady=5, sticky="w")

#Seleção do arquivo de empresa
ttk.Label(frame_campos, text="Selecione a Razão Social:",).grid(row=1, column=0, padx=5, pady=5, sticky="w")
combo_razao_social = ttk.Combobox(frame_campos, textvariable=razao_social_var, width=40)
combo_razao_social.set('Carregar o arquivo de Empresas')
combo_razao_social.state(['readonly'])
combo_razao_social.grid(row=1, column=1, padx=5, pady=5)

combo_razao_social.bind("<Button-1>", lambda e: combo_razao_social.event_generate("<Down>"))
combo_razao_social.bind("<<ComboboxSelected>>", lambda e: preencher_detalhes())

ttk.Label(frame_campos, text="CNPJ:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
ttk.Entry(frame_campos, textvariable=cnpj_var, state='readonly', width=42).grid(row=2, column=1, columnspan=2, padx=5, pady=5, sticky="w")

ttk.Button(frame_campos, text="Selecionar Arquivo", command=selecionar_arquivo_empresas, width=20, bootstyle="light").grid(row=1, column=3, padx=5, pady=5)

ttk.Label(frame_campos, text="Chave API:").grid(row=3, column=0, padx=5, pady=5, sticky="w")
ttk.Entry(frame_campos, textvariable=chave_var, state='readonly', width=42).grid(row=3, column=1, columnspan=3, padx=5, pady=5, sticky="w")

ttk.Label(frame_campos, text="CPF Responsável:").grid(row=4, column=0, padx=5, pady=5, sticky="w")
ttk.Entry(frame_campos, textvariable=cpf_var, state='readonly', width=42).grid(row=4, column=1, columnspan=3, padx=5, pady=5, sticky="w")

ttk.Button(frame_campos, text="Gravar Info", command=confirmar_selecao, width=20, bootstyle="light").grid(row=5, column=3, padx=10)

botao_frame = ttk.Frame(frame_selecao)
botao_frame.grid(row=7, column=0, pady=15, sticky="ew")

for i in range(4):
    botao_frame.grid_columnconfigure(i, weight=1)

botao_width = 25


#Frame do Log de envio e coleta

frame_log = ttk.Frame(frame_selecao, padding=5)
frame_log.grid(row=8, column=0, pady=10, padx=5, sticky="nsew")

log_widget = tk.Text(frame_log, height=10, wrap="word", state="disabled", relief="solid", borderwidth=1)
log_widget.grid(row=0, column=0, sticky="nsew")

scrollbar = ttk.Scrollbar(frame_log, orient="vertical", command=log_widget.yview)
scrollbar.grid(row=0, column=1, sticky="ns")
log_widget.configure(yscrollcommand=scrollbar.set)

frame_log.grid_columnconfigure(0, weight=1)
frame_log.grid_rowconfigure(0, weight=1)

#Loop para rodar a interface
alterar_tema("Kairos")

root.mainloop()
